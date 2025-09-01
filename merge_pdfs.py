import tkinter as tk
from tkinter import messagebox, filedialog
from pathlib import Path
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
import PyPDF2
import os
import re
import time
import shutil
import openpyxl
from PIL import Image
import pytesseract
import fitz
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import math
# import pydicom
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from datetime import datetime

import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="PyPDF2._cmap")
# This is the non working code that is still not completed , which needs a lot of additional work to be done on it.
# def merge_redcliffe_pdf_files():
#     input_directory = filedialog.askdirectory(title="Select Input Directory")

#     if input_directory:
#         output_directory = filedialog.askdirectory(title="Select Output Directory")

#         if output_directory:
#             pdf_dir = Path(input_directory)
#             pdf_output_dir = Path(output_directory)
#             merged_files_dir = pdf_output_dir / "Merged Files"
#             merged_files_dir.mkdir(parents=True, exist_ok=True)

#             pdf_files = list(pdf_dir.glob("*.pdf"))
#             print("These are the pdf files list :", pdf_files)

#             naming_errors = {}
#             exception_files = {}
#             processed_ids = set()
#             keys = set()
#             merged_count = 0

#             for file in pdf_files:
#                 try:
#                     file_id = str(file).split("\\")[-1].split("_")[0].lower()
#                     keys.add(file_id)
#                 except IndexError:
#                     original_filename = str(file).split("\\")[-1]
#                     naming_errors[str(file)] = original_filename
#                     keys.add(file_id)
#                     print(f"File {file} has incorrect naming format. Storing naming error: {original_filename}")

#             for key in keys:
#                 if key in processed_ids:
#                     print(f"Skipping all files with ID: {key} due to already processed.")
#                     continue

#                 xray, optometry, ecg_report, pft, audiometry, vitals, blood_report, smart_report, vaccination_report = [None] * 9
#                 others = []
#                 print("This is the key :", key)

#                 for file in pdf_files:
#                     str_pdf_file = str(file)
#                     print(str_pdf_file)  # Log added
#                     splitted_file = str_pdf_file.rsplit("\\", 1)[1].lower()
#                     print("This is the splitted file :", splitted_file)  # Log added
#                     try:
#                         file_id = splitted_file.split("_")[0].lower()
#                     except IndexError:
#                         continue
#                     if file_id != key:
#                         continue

#                     try:
#                         print("Inside the try block.")  # Log added
#                         with open(file, "rb") as f:
#                             pdf_reader = PdfReader(f)
#                             print("After reading the pdf.")  # Log added
#                             first_page_text = pdf_reader.pages[0].extract_text() if len(pdf_reader.pages) >= 1 else ""
#                             second_page_text = pdf_reader.pages[1].extract_text() if len(pdf_reader.pages) >= 2 else ""
#                             print("This is the first page text :")  # Log added
#                             print(first_page_text)  # Log added
#                             print("End of first page text")  # Log added
#                             print("This is the second page text:")  # Log added
#                             print(second_page_text)  # Log added
#                             print("End of second page text")  # Log added

#                             if "X-RAY" in first_page_text or "X-RAY" in second_page_text:
#                                 xray = file
#                                 print("This is an xray file.")  # Log added
#                             elif "OPTOMETRY" in first_page_text:
#                                 optometry = file
#                                 print("This is an opto file.")  # Log added
#                             elif "ECG" in second_page_text or "Acquired on" in second_page_text:
#                                 ecg_report = file
#                                 print("This is an ecg file.")  # Log added
#                             elif "RECORDERS & MEDICARE SYSTEMS" in first_page_text:
#                                 pft = file
#                                 print("This is a pft file.")  # Log added
#                             elif "VITALS" in first_page_text:
#                                 vitals = file
#                                 print("This is a vitals file.")  # Log added
#                             elif 'left ear' in first_page_text:
#                                 audiometry = file
#                                 print("This is a audio file.")  # Log added
#                             elif 'RBC Count' in first_page_text or "PDW *" in second_page_text or "PDW" in second_page_text:
#                                 blood_report = file
#                                 print("This is a blood report.")  # Log added
#                             elif 'SMART REPORT' in first_page_text:
#                                 smart_report = file
#                                 print("This is a smart report.")  # Log added
#                             elif 'CERTIFICATE OF VACCINATION' in first_page_text:
#                                 vaccination_report = file
#                                 print("This is a vaccination file.")  # Log added
#                             else:
#                                 others.append(file)
#                                 print("This is an others file.")  # Log added

#                     except Exception as e:
#                         print(f"Error processing file: {file}")  # Log added
#                         original_filename = str(file).split("\\")[-1]
#                         exception_files[str(file)] = (original_filename, str(e))
#                         print(f"This file has an exception : {original_filename}")  # Log added
#                         continue

#                 processed_ids.add(key)
#                 if any([xray, optometry, ecg_report, pft, audiometry, vitals, others, blood_report, smart_report, vaccination_report]):
#                     merger = PdfMerger()

#                     if smart_report: merger.append(smart_report)
#                     if vitals: merger.append(vitals)
#                     if ecg_report: merger.append(ecg_report)
#                     if pft: merger.append(pft)
#                     if audiometry: merger.append(audiometry)
#                     if optometry: merger.append(optometry)
#                     if xray: merger.append(xray)
#                     if blood_report: merger.append(blood_report)
#                     if vaccination_report: merger.append(vaccination_report)
#                     c

#                     base_file_name = (
#                         xray.stem.split(".")[0].lower() if xray else
#                         optometry.stem.split(".")[0].lower() if optometry else
#                         ecg_report.stem.split(".")[0].lower() if ecg_report else
#                         pft.stem.split(".")[0].lower() if pft else
#                         audiometry.stem.split(".")[0].lower() if audiometry else
#                         vitals.stem.split(".")[0].lower() if vitals else
#                         vaccination_report.stem.split(".")[0].lower() if vaccination_report else
#                         smart_report.stem.split(".")[0].lower() if smart_report else
#                         blood_report.stem.split(".")[0].lower() if blood_report else
#                         "NonRecognizable"
#                     )
#                     merged_file_path = merged_files_dir / f"{base_file_name}.pdf"
#                     merger.write(str(merged_file_path))
#                     print(f"Merged PDF saved to: {merged_file_path}")
#                     merged_count += 1

#             if exception_files:
#                 problematic_files_dir = pdf_output_dir / "Problematic Files"
#                 problematic_files_dir.mkdir(parents=True, exist_ok=True)
#                 for file_path, (original_name, exception) in exception_files.items():
#                     try:
#                         shutil.copy2(file_path, problematic_files_dir / original_name)
#                         print(f"Copied problematic file: {file_path} to {problematic_files_dir / original_name}")
#                     except Exception as e:
#                         print(f"Error copying problematic file: {file_path}. Error: {e}")

#             total_input_count = len(pdf_files)

#             if naming_errors:
#                 problem_files_str = "\n".join([f"{i+1}. {name}" for i, name in enumerate(naming_errors.values())])
#                 rename_issue_files_count = len(naming_errors)
#                 tk.messagebox.showwarning("Naming Errors",
#                                         f"Total {rename_issue_files_count} files have incorrect naming:\n{problem_files_str}\n\nNOTE:\nThese files were still merged if their IDs matched other correctly named files.")

#             if exception_files:
#                 exception_files_str = "\n".join([f"File : {name} \n Exception : {exception}" for name, (name, exception) in exception_files.values()])
#                 problematic_file_count = len(exception_files)
#                 tk.messagebox.showwarning("File Exceptions",
#                                         f"Total {problematic_file_count} files had this issue : \n {exception_files_str} \n If in case you are not able to solve this problem, Contact Himanshu .")


#             tk.messagebox.showinfo("PDF Merger", f"Total {total_input_count} PDF files processed. {merged_count} PDF files were merged successfully!")

#         else:
#             tk.messagebox.showwarning("Output Directory", "Output directory not selected.")
#     else:
#         tk.messagebox.showwarning("Input Directory", "Input directory not selected.")



# This is my code working till version 4 , and all the changes i've done is mentioned in the readme file - Himanshu. 
def merge_redcliffe_pdf_files():
    # Prompt user to select input directory
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        # Prompt user to select output directory
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            pdf_dir = Path(input_directory)
            pdf_output_dir = Path(output_directory)
            pdf_output_dir.mkdir(parents=True, exist_ok=True)

            # List all PDF files in the input directory
            pdf_files = list(pdf_dir.glob("*.pdf"))
            print("These are the pdf files list :", pdf_files)

            # keys = set([str(file).split("\\")[-1].split("_")[0].lower() for file in pdf_files])
            keys = set()
            naming_errors = {}
            exception_files = {}

            for file in pdf_files:
                # This code is explicitly written to check the name format of the files - Himanshu.
                try:
                    original_filename = str(file).split("\\")[-1]
                    file_id = str(file).split("\\")[-1].split("_")[0].lower()
                    if "." in file_id:
                        naming_errors[str(file)] = original_filename
                        print(f"File {file} has incorrect naming format. Storing naming error: {original_filename}")
                    else:
                        keys.add(file_id)
                except IndexError:
                    original_filename = str(file).split("\\")[-1]
                    naming_errors[str(file)] = original_filename
                    # keys.add(file_id)
                    print(f"File {file} has incorrect naming format. Storing naming error: {original_filename}")



            for key in keys:
                xray, optometry, ecg_report, pft, audiometry, vitals, blood_report, smart_report, vaccination_report = [None] * 9
                others = []
                # pdf_files_3_or_less = False
                # total_pdfs_less_than_3 = 0

                print("This is the key :", key)

                for file in pdf_files:
                    print("Entering the for files in pdf_files loop")
                    str_pdf_file = str(file)
                    print(str_pdf_file)
                    # split_str_pdf_files = str_pdf_file.split("_")[1].lower()
                    # print("This is the splitted str :", split_str_pdf_files)
                    splitted_file = str_pdf_file.rsplit("\\",1)[1].lower()
                    print("This is the splitted file :", splitted_file)
                    try:
                        print("Inside the try block.")
                        # if splitted_file.startswith(key):
                        #     print("Starts with key")
                        # if split_str_pdf_files.endswith(key):
                        if splitted_file.startswith(key):
                            print("Starts with key")
                            first_page_text = ''
                            second_page_text = ''
                            print("Before opening the pdf file page.")
                            pdf_reader = PdfReader(open(file, "rb"))
                            print("After reading the pdf.")
                            if len(pdf_reader.pages) == 1:
                                print("if the pdf reader length is equal to 1")
                                first_page = pdf_reader.pages[0]
                                first_page_text = first_page.extract_text()
                            else:
                                print("this is the else statement after calculating the pdf length.")
                                second_page = pdf_reader.pages[1]
                                second_page_text = second_page.extract_text()

                            print("This is the first page text :")
                            print(first_page_text)
                            print("End of first page text")
                            print("This is the second page text:")
                            print(second_page_text)
                            print("End of second page text")

                            if "X-RAY" in first_page_text or "X-RAY" in second_page_text:
                                xray = file
                                print("This is an xray file.")
                            elif "OPTOMETRY" in first_page_text:
                                optometry = file
                                print("This is an opto file.")
                            elif "ECG" in second_page_text or "Acquired on" in second_page_text:
                                ecg_report = file
                                print("This is an ecg file.")
                            elif "RECORDERS & MEDICARE SYSTEMS" in first_page_text:
                                pft = file
                                print("This is a pft file.")
                            elif "VITALS" in first_page_text:
                                vitals = file
                                print("This is a vitals file.")
                            elif 'left ear' in first_page_text:
                                audiometry = file
                                print("This is a audio file.")
                            elif 'RBC Count' in first_page_text or "PDW *" in second_page_text or "PDW" in second_page_text:
                                blood_report = file
                                print("This is a blood report.")
                            elif 'SMART REPORT' in first_page_text:
                                smart_report = file
                                print("This is a smart report.")
                            elif 'CERTIFICATE OF VACCINATION' in first_page_text:
                                vaccination_report = file
                                print("This is a vaccination file.")
                            else:
                                others.append(file)
                                print("This is an others file.")


                    except Exception as e:
                        print(f"Error processing file: {file}")
                        print(f"Error details: {str(e)}")

                        # Move the problematic file to the error folder
                        # error_folder = pdf_output_dir / "Problematic Files"
                        # error_folder.mkdir(parents=True, exist_ok=True)
                        # # move_to_error_folder(file, error_folder)
                        # shutil.copy2(file, error_folder)
                        # print(f"This is the problematic file : {file}")
                        original_filename = str(file).split("\\")[-1]
                        exception_files[original_filename] = str(e)  # Store in dictionary: filename as key, error as value
                        print(f"This is the problematic file : {file}")

                # Check if at least one file is available for merging
                # if xray or optometry or ecg_report or pft or audiometry or vitals or others:
                if any([xray, optometry, ecg_report, pft, audiometry, vitals, others, blood_report, smart_report, vaccination_report]):
                    merger = PdfMerger()
                    if smart_report:
                        merger.append(smart_report)
                    if vitals:
                        merger.append(vitals)
                    if ecg_report:
                        merger.append(ecg_report)
                    if pft:
                        merger.append(pft)
                    if audiometry:
                        merger.append(audiometry)
                    if optometry:
                        merger.append(optometry)
                    if xray:
                        merger.append(xray)
                    if blood_report:
                        merger.append(blood_report)
                    if vaccination_report:
                        merger.append(vaccination_report)
                    # if others:
                    #     merger.append(others)
                    for other_file in others:
                        merger.append(other_file)

                    # if len(merger.pages) >= 1:
                    #     merged_pdf_dir = pdf_output_dir / "7_pages"
                    # else:
                    #     pass

                    base_file_name = (
                        xray.stem.split(".")[0].lower() if xray else
                        optometry.stem.split(".")[0].lower() if optometry else
                        ecg_report.stem.split(".")[0].lower() if ecg_report else
                        pft.stem.split(".")[0].lower() if pft else
                        audiometry.stem.split(".")[0].lower() if audiometry else
                        vitals.stem.split(".")[0].lower() if vitals else
                        vaccination_report.stem.split(".")[0].lower() if vaccination_report else
                        others[0].stem.split(".")[0].lower() if others else
                        "NonRecognizable"
                    )
                    print(base_file_name)

                    merged_pdf_dir = pdf_output_dir / "Merged Files"
                    merged_file_path = merged_pdf_dir / f"{base_file_name}.pdf"
                    merged_file_path.parent.mkdir(parents=True, exist_ok=True)

                    merger.write(str(merged_file_path))
                    print(f"Merged PDF saved to: {merged_file_path}")

            # Display message box after merging is complete
            total_input_count = len(pdf_files)
            final_pagers = list((pdf_output_dir / "Merged Files").glob("*.pdf"))
            total_pdfs = len(final_pagers)

            total_count =  total_pdfs
            # Adding the count logics :
            # if total_count <= 3 :
            #     pdf_files_3_or_less = True
            #     # updating the count.
            #     total_pdfs_less_than_3 += 1

            tk.messagebox.showinfo("PDF Merger", f"Total {total_input_count} PDF files merged into {total_count} PDF files successfully!")

            # Display message box with Less than 3 Pages if any
            # if pdf_files_3_or_less:
            #     # file_list = "\n".join(str(file) for file in pdf_files_3_or_less)
            #     tk.messagebox.showinfo("Missing Files",
            #                            f"Total {total_pdfs_less_than_3} merged PDF files have only less then 3 or more then 4 pages.")
            # else:
            #     tk.messagebox.showinfo("No Missing Files", "All merged PDF files have 3 or 4 pages.")

            if naming_errors:
                # Write naming errors to "NamingErrors.txt" under "Naming Conflict"
                naming_conflict_dir = pdf_output_dir / "Naming Conflict"
                naming_conflict_dir.mkdir(parents=True, exist_ok=True)

                naming_errors_path = naming_conflict_dir / "NamingErrors.txt"
                with open(naming_errors_path, "w") as file:
                    file.write(f"{len(naming_errors)} files are having naming conflicts:\n----------------------------------------------------------------------\n")
                    # Enumerate to add index (starting from 1)
                    for index, filename in enumerate(naming_errors.values(), 1):
                        file.write(f"{index}. {filename}\n")

                naming_error_messages = []  # List to store formatted naming error messages
                for file_path, original_name in naming_errors.items():
                    naming_error_messages.append(original_name)

                naming_errors_str = "\n".join(naming_error_messages) #joining the naming error message with new line.
                rename_issue_files_count = len(naming_errors)
                tk.messagebox.showwarning("Naming Errors",
                                        f"{rename_issue_files_count} file(s) is/are having the naming errors:\n\n{naming_errors_str}\n\nNOTE:\nThese files might've still got merged if their IDs matched other correctly named files, PLEASE CHECK!")

            if exception_files:  # Only if there are exception files
                error_folder = pdf_output_dir / "Problematic Files"
                error_folder.mkdir(parents=True, exist_ok=True)

                error_messages = []  # List to store formatted error messages
                for filename, error_message in exception_files.items():
                    try:
                        original_file_path = pdf_dir / filename
                        shutil.copy2(original_file_path, error_folder / filename)  # Copy with original filename
                        print(f"Copied problematic file: {original_file_path} to {error_folder / filename}")
                        error_messages.append(f"Name: {filename}, Error: {error_message}") #appending the error message
                    except FileNotFoundError:
                        print(f"Original file not found: {original_file_path}. Cannot copy.")
                        error_messages.append(f"Name: {filename}, Error: File Not Found") #appending the error message
                    except Exception as copy_error:
                        print(f"Error copying problematic file: {original_file_path}. Error: {copy_error}")
                        error_messages.append(f"Name: {filename}, Error: {copy_error}") #appending the error message

                #Create the final message string
                final_error_message = "These are the list of files that had errors:\n" + "\n".join(error_messages) + "\nIf problem is unrecognizable, Contact Himanshu."
                print(final_error_message) #printing the final error message
                tk.messagebox.showwarning("File Exceptions", final_error_message)
        else:
            tk.messagebox.showwarning("Output Directory", "Output directory not selected.")
    else:
        tk.messagebox.showwarning("Input Directory", "Input directory not selected.")


# def merge_redcliffe_pdf_files():
#     # Prompt user to select input directory
#     input_directory = filedialog.askdirectory(title="Select Input Directory")
#     if not input_directory:
#         tk.messagebox.showwarning("Input Directory", "Input directory not selected.")
#         return

#     # Prompt user to select output directory
#     output_directory = filedialog.askdirectory(title="Select Output Directory")
#     if not output_directory:
#         tk.messagebox.showwarning("Output Directory", "Output directory not selected.")
#         return

#     pdf_dir = Path(input_directory)
#     pdf_output_dir = Path(output_directory)
#     pdf_output_dir.mkdir(parents=True, exist_ok=True)

#     # List all PDF files in the input directory
#     pdf_files = list(pdf_dir.glob("*.pdf"))

#     # Dictionary to categorize files by patient ID
#     patient_files = {}

#     for file in pdf_files:
#         # Extract patient ID from filename
#         parts = file.stem.split("_")
#         patient_id = parts[0].lower()  # Use lowercase for consistency

#         if patient_id not in patient_files:
#             patient_files[patient_id] = {'xray': None, 'optometry': None, 'ecg': None, 'pft': None, 'audiometry': None, 'vitals': None, 'others': None}

#         try:
#             pdf_reader = PdfReader(open(file, "rb"))
#             first_page_text = pdf_reader.pages[0].extract_text() if len(pdf_reader.pages) > 0 else ''
#             second_page_text = pdf_reader.pages[1].extract_text() if len(pdf_reader.pages) > 1 else ''
            
#             print(f"Processing file: {file}")
#             print(f"First Page Text: {first_page_text[:100]}")  # Print first 100 chars for debugging
#             print(f"Second Page Text: {second_page_text[:100]}")  # Print first 100 chars for debugging

#             if "Study Date" in first_page_text and "Report Date" in first_page_text:
#                 patient_files[patient_id]['xray'] = file
#             elif "OPTOMETRY" in first_page_text:
#                 patient_files[patient_id]['optometry'] = file
#             elif "ECG" in second_page_text:
#                 patient_files[patient_id]['ecg'] = file
#             elif "RECORDERS & MEDICARE SYSTEMS" in first_page_text:
#                 patient_files[patient_id]['pft'] = file
#             elif "VITALS" in first_page_text:
#                 patient_files[patient_id]['vitals'] = file
#             elif 'left ear' in first_page_text:
#                 patient_files[patient_id]['audiometry'] = file
#             else:
#                 patient_files[patient_id]['others'] = file

#         except Exception as e:
#             print(f"Error processing file: {file}")
#             print(f"Error details: {str(e)}")
#             # Move the problematic file to the error folder
#             error_folder = pdf_output_dir / "error_pdf"
#             error_folder.mkdir(parents=True, exist_ok=True)
#             move_to_error_folder(file, error_folder)

#     # Process each patient and merge PDFs
#     total_input_count = len(pdf_files)
#     merged_pdfs_count = 0

#     for patient_id, files in patient_files.items():
#         print(f"Merging files for patient ID: {patient_id}")
#         merger = PdfMerger()
#         for category in ['vitals', 'ecg', 'xray', 'pft', 'audiometry', 'optometry', 'others']:
#             if files[category]:
#                 print(f"Appending file: {files[category]}")
#                 merger.append(files[category])
        
#         if len(merger.pages) > 0:
#             merged_pdf_dir = pdf_output_dir / "7_pages"
#             merged_pdf_dir.mkdir(parents=True, exist_ok=True)
            
#             base_file_name = patient_id
#             merged_file_path = merged_pdf_dir / f"{base_file_name}.pdf"
#             merger.write(str(merged_file_path))
#             merger.close()
#             merged_pdfs_count += 1
#             print(f"Merged PDF saved to: {merged_file_path}")

#     # Display message box after merging is complete
#     tk.messagebox.showinfo("PDF Merger", f"Total {total_input_count} PDF files processed. {merged_pdfs_count} PDFs merged successfully!")

# def move_to_error_folder(file, error_folder):
#     try:
#         shutil.move(str(file), error_folder / file.name)
#     except Exception as e:
#         print(f"Error moving file to error folder: {file}")
#         print(f"Error details: {str(e)}")

# if __name__ == "__main__":
#     merge_redcliffe_pdf_files()

def merge_all():
    # Prompt user to select input directory
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        # Prompt user to select output directory
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            pdf_dir = Path(input_directory)
            pdf_output_dir = Path(output_directory)
            pdf_output_dir.mkdir(parents=True, exist_ok=True)

            # List all PDF files in the input directory
            pdf_files = list(pdf_dir.glob("*.pdf"))
            if pdf_files:
                # Create a PdfFileMerger object
                merger = PdfMerger()
                for pdf_file in pdf_files:
                    merger.append(pdf_file)
                output_file_path = pdf_output_dir / "merged_file.pdf"

                # Write the merged PDF to the output file
                with open(output_file_path, "wb") as output_file:
                    merger.write(output_file)

                print(f"Merged PDF saved to: {output_file_path}")
                tk.messagebox.showinfo("PDF Merger",f"Total {len(pdf_files)} PDF files merged into one PDF successfully!")
            else:
                tk.messagebox.showinfo("No PDF Files", "No PDF files found in the input directory.")
        else:
            tk.messagebox.showwarning("Output Directory", "Output directory not selected.")
    else:
        tk.messagebox.showwarning("Input Directory", "Input directory not selected.")

import shutil
import re
from tkinter import filedialog, messagebox
from pathlib import Path
import PyPDF2

def rename_pdf_files():
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            input_dir = Path(input_directory)
            output_dir = Path(output_directory)
            output_dir.mkdir(parents=True, exist_ok=True)

            error_dir = output_dir / "error_files"
            error_dir.mkdir(parents=True, exist_ok=True)

            pdf_files = list(input_dir.glob("*.pdf"))

            if pdf_files:
                renamed_count = 0
                error_count = 0

                for pdf_file in pdf_files:
                    try:
                        with open(pdf_file, 'rb') as file:
                            pdf_reader = PyPDF2.PdfReader(file)
                            if len(pdf_reader.pages) == 0:
                                raise ValueError("PDF has no pages")

                            # Extract text from first or second page
                            first_page = pdf_reader.pages[0]
                            first_page_text = first_page.extract_text()

                            second_page_text = ""
                            if len(pdf_reader.pages) > 1:
                                second_page_text = pdf_reader.pages[1].extract_text()

                            patient_id = ""
                            patient_name = ""

                            # ---- BLOOD REPORT LOGIC ----
                            if "RBC Count" in first_page_text:
                                if "Patient Name :" in first_page_text:
                                    complete_name = first_page_text.split("Patient Name : ")[1].split("DOB/")[0].strip()
                                    patient_id = complete_name.rsplit(" ", 1)[1]
                                    patient_name = complete_name.rsplit(" ", 1)[0].split(" ", 1)[1].lower()
                                elif "Patient NAME :" in first_page_text:
                                    complete_name = first_page_text.split("Patient NAME : ")[1].split("DOB/")[0].strip()
                                    patient_id = complete_name.rsplit("_", 1)[1]
                                    patient_name = complete_name.rsplit("_", 1)[0].split(" ", 1)[1].lower()

                            # ---- X-RAY REPORT ----
                            elif "Study Date" in first_page_text and "Report Date" in first_page_text:
                                patient_id = first_page_text.split("Patient ID")[1].split(" ")[1].strip().lower()
                                patient = first_page_text.split("Name")[1].split("Date")[0].split(" ")[0].strip().lower()
                                patient_name = patient.split("patient")[0].strip() if "patient" in patient else patient

                            # ---- PFT REPORT (your sample PDF) ----
                            elif "RECORDERS" in first_page_text:
                                match_id = re.search(r"ID\s*:?\s*([A-Za-z0-9]+)", first_page_text)
                                patient_id = match_id.group(1).lower() if match_id else "unknown"

                                match_name = re.search(r"Patient:\s*([^\n\r]+)", first_page_text)
                                patient_name = match_name.group(1).strip().split()[0].lower() if match_name else "unknown"

                            # ---- ECG GRAPH ----
                            elif "Acquired on:" in first_page_text:
                                if "Id :" in first_page_text:
                                    patient_id = first_page_text.split("Id :")[1].split(" ")[1].split("\n")[0].strip().lower()
                                elif "Id:" in first_page_text:
                                    patient_id = first_page_text.split("Id:")[1].split(" ")[1].split("\n")[0].strip().lower()

                                if "Name :" in first_page_text:
                                    patient_name = first_page_text.split("Name :")[1].split("Age")[0].split(" ")[1].strip().lower()
                                elif "Name:" in first_page_text:
                                    patient_name = first_page_text.split("Name:")[1].split("Age")[0].split(" ")[1].strip().lower()

                            # ---- ECG from second page ----
                            elif "ECG" in second_page_text:
                                patient_id = second_page_text.split("Patient ID:")[1].split("Age:")[0].strip().lower()
                                patient_name = second_page_text.split("Name:")[1].split("Patient ID:")[0].strip().lower()

                            # ---- Audiometry ----
                            elif "left ear" in first_page_text:
                                patient_id = first_page_text.split('Patient ID')[1].split('Age')[0].strip().lower()
                                patient_name = first_page_text.split('Name')[1].split('Patient ID')[0].strip().lower()

                            # ---- Optometry & Vitals ----
                            elif "OPTOMETRY" in first_page_text or "VITALS" in first_page_text:
                                patient_id = first_page_text.split("Patient ID:")[1].split("Age:")[0].strip().lower()
                                patient_name = first_page_text.split("Name:")[1].split("Patient ID:")[0].strip().lower()

                            # Validate extracted values
                            if not patient_id or not patient_name or patient_id == "unknown" or patient_name == "unknown":
                                raise ValueError("Missing or invalid patient details")

                            # Rename file
                            new_filename = f"{patient_id}_{patient_name}{pdf_file.suffix}"
                            new_file_path = output_dir / new_filename
                            shutil.copy2(pdf_file, new_file_path)
                            renamed_count += 1
                            print(f"✔ Renamed: {pdf_file.name} -> {new_filename}")

                    except Exception as e:
                        error_count += 1
                        error_file_path = error_dir / pdf_file.name
                        shutil.copy2(pdf_file, error_file_path)
                        print(f"❌ Error processing {pdf_file.name}: {str(e)}")

                # Summary
                messagebox.showinfo("Renaming Complete", f"{renamed_count} PDF files have been renamed.")
                if error_count > 0:
                    messagebox.showwarning("Errors Encountered", f"{error_count} PDF files had issues and were saved in 'error_files'.")

            else:
                messagebox.showwarning("No PDFs Found", "No PDF files found in the selected input directory.")
        else:
            messagebox.showwarning("Missing Output Directory", "You must select an output directory.")
    else:
        messagebox.showwarning("Missing Input Directory", "You must select an input directory.")

def remove_illegal_characters(value):
    if isinstance(value, str):
        # Remove characters that are not printable or are control characters
        value = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', value)
    return value

# This function extract data from individual files (not merged ones) i.e. i have to pass all the folder where all the files are present,
# i.e. all the files of all patients, then one by one all the data will get appended to the excel sheet. - HIMANSHU.
def generate_excel_for_individual_files():
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            input_dir = Path(input_directory)
            output_dir = Path(output_directory)
            output_dir.mkdir(parents=True, exist_ok=True)

            error_dir = output_dir / "error_files"
            error_dir.mkdir(parents=True, exist_ok=True)

            pdf_files = list(input_dir.glob("*.pdf"))
            error_count = 0

            # Define unwanted phrases
            unwanted_phrases = ["Referrer Dr.", "left ear", "RECORDERS & MEDICARE SYSTEMS", "OPTOMETRY", "ECG", "VITALS", "RBC Count", "PDW *", "PDW"]

            # There are mainly 8 files here (leaving multiple formats of an individual) :
            patient_data_ecg = []
            patient_data_ecg1 = []
            patient_data_pft = []
            patient_data_xray = []
            # Adding the remaining files too.
            patient_data_opto = []
            patient_data_vitals = []
            patient_data_audio = []
            patient_data_blood = []

            # These are the count of individual files :
            # total_ecg_files, total_ecg_files1, total_pft_files, total_xray_files, total_opto_files, total_vitals_files, total_audio_files, total_blood_files = [] * 8
            total_ecg_files = 0
            total_ecg_files1 = 0
            total_pft_files = 0
            total_xray_files = 0
            total_opto_files = 0
            total_vitals_files = 0
            total_audio_files = 0
            total_blood_files = 0

            # This is the excel file path for all these files :
            # excel_file_path_ecg, excel_file_path_pft, excel_file_path_xray, excel_file_path_opto, excel_file_path_vitals, excel_file_path_audio, excel_file_path_blood = "" * 7
            excel_file_path_ecg = ""
            excel_file_path_pft = ""
            excel_file_path_xray = ""
            excel_file_path_opto = ""
            excel_file_path_vitals = ""
            excel_file_path_audio = ""
            excel_file_path_blood = ""

            # Above i don't need to make the path for ecg1(i.e. coming for the drive.), because it is going to be appended to the place of 'ecg' only.

            workbook_xray = Workbook()
            sheet_xray = workbook_xray.active
            row_xray = 2

            for pdf_file in pdf_files:
                with open(pdf_file, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    if len(pdf_reader.pages) > 0:
                        first_page_text = ''
                        second_page_text = ''
                        if len(pdf_reader.pages) == 1:
                            first_page = pdf_reader.pages[0]
                            first_page_text = first_page.extract_text()
                        elif len(pdf_reader.pages) == 2:
                            first_page = pdf_reader.pages[1]
                            first_page_text = first_page.extract_text()
                        # Fixing this thing for our pathology / blood data extraction.
                        elif len(pdf_reader.pages) > 2:
                            first_page = pdf_reader.pages[0]
                            first_page_text = first_page.extract_text()
                        else:
                            second_page = pdf_reader.pages[1]
                            second_page_text = second_page.extract_text()
                            first_page_text = second_page_text

                        try:
                            # # xray
                            # if "Study Date" in first_page_text or "Report Date" in first_page_text:
                            #     patient_id = str(first_page_text).split("Patient ID")[1].split(" ")[1].lower().strip()
                            #     patient = str(first_page_text).split("Name")[1].split("Date")[0].split(" ")[0].strip().lower()
                            #     if "patient" in patient:
                            #         patient_name = patient.split("patient")[0].strip()
                            #     else:
                            #         patient_name = patient
                            #     gender = str(first_page_text).split("Sex")[1].split("Study Date")[0].strip().lower()
                            #     if 'Yr' or 'Y' or 'yrs' in first_page_text:
                            #         if 'Yr' in first_page_text:
                            #             age_data = str(first_page_text).split("Age")[1].split("Yr")[0].strip()
                            #             if "Days" in age_data:
                            #                 age = age_data.split("Days")[0]
                            #             else:
                            #                 age = age_data
                            #         if 'Y' in first_page_text:
                            #             age_data = str(first_page_text).split("Age")[1].split('Y')[0].strip()
                            #             if "Days" in age_data:
                            #                 age = age_data.split("Days")[0]
                            #             else:
                            #                 age = age_data
                            #         if 'yrs' in first_page_text:
                            #             age_data = str(first_page_text).split("Age")[1].split('yrs')[0].strip()
                            #             if "Days" in age_data:
                            #                 age = age_data.split("Days")[0]
                            #             else:
                            #                 age = age_data

                            #     test_date = str(first_page_text).split("Study Date")[1].split("\n")[1].split("Time")[1]
                            #     report_date = str(first_page_text).split("Report Date")[1].split("\n")[1].split("Time")[1]

                            #     if "Adv: Clinical correlation." not in first_page_text:
                            #         findings_data = str(first_page_text).split("IMPRESSION")[1].split("Correlate clinically")[0].split(":")[1].strip()
                            #         if "Please" in findings_data:
                            #             findings_with_dot = findings_data.split("Please")[0]
                            #             if "•" in findings_with_dot:
                            #                 findings = findings_with_dot.split("•")[1].split(".")[0]
                            #             else:
                            #                 findings = findings_with_dot.split(".")[0]
                            #         else:
                            #             findings_with_dot = findings_data
                            #             if "•" in findings_with_dot:
                            #                 findings = findings_with_dot.split("•")[1].split(".")[0]
                            #             else:
                            #                 findings = findings_with_dot.split(".")[0]


                            #     if "Adv: Clinical correlation." in first_page_text:
                            #         findings_data1 = str(first_page_text).split("Impression")[1]
                            #         if findings_data1:
                            #             findings = findings_data1.split("Adv: Clinical correlation.")[0].split(':')[1].strip()


                            #     if  findings == 'No significant abnormality noted' or findings == 'No significant abnormality':
                            #         findings = 'No significant abnormality seen'
                            #     patient_data_xray.append((patient_id, patient_name, age, gender, test_date, report_date, remove_illegal_characters(findings)))
                            #     print(patient_id, patient_name, age, gender, test_date, report_date, findings)
                            #     total_xray_files += 1

                            # Printing the first page text data everytime it's processed.
                            print(f"This is the extracted text data : {first_page_text}")

                            # For the xray report from u4rad pacs reporting bot.
                            if "Test Date:" in first_page_text and "Report Date:" in first_page_text and not any(phrase in first_page_text for phrase in unwanted_phrases):
                                print("This is an xray report from our u4rad pacs.")
                                # patient_data_xray = extract_data_from_the_u4rad_pacs_xray_file(first_page_text)
                                # print(patient_data_xray)
                                # if isinstance(patient_data_xray, dict):  # If it's a dictionary, convert it into a tuple or list
                                #     patient_data_xray = tuple(patient_data_xray.values())
                                #     print(f"Data after making it a tuple : {patient_data_xray}")
                                # # else:
                                # #     data = patient_data_xray
                                # total_xray_files += 1
                                data = extract_data_from_the_u4rad_pacs_xray_file(first_page_text)
                                if isinstance(data, dict):
                                    data_tuple = tuple(data.values())
                                elif isinstance(data, tuple):
                                    data_tuple = data
                                else:
                                    data_tuple = (data,)  # Handle other cases if needed
                                patient_data_xray.append(data_tuple)
                                total_xray_files += 1
                                print(f"Data added: {data_tuple}")

                            # Extract ECG data
                            elif "Acquired on:" in first_page_text:
                                patient_id = str(first_page_text).split("Id :")[1].split(" ")[1].split("\n")[0]
                                patient_name = str(first_page_text).split("Name :")[1].split("Age :")[0]
                                patient_age = str(first_page_text).split("Age :")[1].split(" ")[1].split("\n")[0]
                                patient_gender = str(first_page_text).split("Gender :")[1].split("|")[0].strip()
                                heart_rate = str(first_page_text).split("HR:")[1].split("R(II):")[0].strip()
                                report_time = str(first_page_text).split("Acquired on:")[1][12:17]
                                report_date = str(first_page_text).split("Acquired on:")[1][1:11]
                                R_II = str(first_page_text).split("R(II):")[1].split("RR")[0].strip()
                                RR = str(first_page_text).split("RR:")[1].split("PR:")[0].strip()
                                PR = str(first_page_text).split("PR:")[1].split("QRS:")[0].strip()
                                QRS = str(first_page_text).split("QRS:")[1].split("QT:")[0].strip()
                                QT = str(first_page_text).split("QT:")[1].split("QTc:")[0].strip()
                                QTc = str(first_page_text).split("QTc:")[1].split("QT/QTc:")[0].split("QT/")[0].strip()
                                QT_QTc = str(first_page_text).split("QT/QTc:")[1].strip()

                                print(patient_id, patient_name, patient_age, patient_gender, heart_rate, report_time, report_date)
                                patient_data_ecg.append((patient_id, patient_name, patient_age, patient_gender, heart_rate,report_time, report_date, R_II
                                                         ,RR, PR, QRS, QT, QTc, QT_QTc))
                                total_ecg_files += 1

                            #pft
                            elif "RECORDERS & MEDICARE SYSTEMS" in first_page_text:
                                patient_id = str(first_page_text).split("ID")[1].split("Age")[0].split(":")[1].strip()
                                patient_name = str(first_page_text).split("Patient")[1].split("Refd.By:")[0].split(":")[1].strip()
                                patient_age = str(first_page_text).split("Age    :")[1].split("Yrs")[0].strip()
                                gender = str(first_page_text).split("Gender")[1].split("Smoker")[0].split(":")[1].strip()
                                height = str(first_page_text).split("Height :")[1].split("Weight")[0].strip()
                                weight = str(first_page_text).split("Weight")[1].split("Gender")[0].split(":")[1].split("Kgs")[0].strip()
                                date = str(first_page_text).split("Date")[1][1:21].split(":")[1]
                                observation = str(first_page_text).split("Pre Test COPD Severity")[1]
                                patient_data_pft.append((patient_id, patient_name, patient_age, gender, height, weight, date, remove_illegal_characters(observation)))
                                print(patient_id, patient_name, patient_age, gender, height, weight, date, observation)
                                total_pft_files += 1
                            
                            # For pathology reports.
                            elif "RBC Count" in first_page_text or "PDW *" in first_page_text or "PDW" in first_page_text :
                                print("This is an blood report from our u4rad redcliffe pathology.")
                                # patient_data_blood = extract_data_from_the_redcliffe_patho_file(first_page_text)
                                # print(patient_data_blood)
                                # if isinstance(patient_data_blood, dict):  # If it's a dictionary, convert it into a tuple or list
                                #     patient_data_blood = tuple(patient_data_blood.values())
                                #     print(f"Data after making it a tuple : {patient_data_blood}")
                                # # else:
                                # #     data = patient_data_blood
                                # total_blood_files += 1
                                data = extract_data_from_the_redcliffe_patho_file(first_page_text)
                                if isinstance(data, dict):
                                    data_tuple = tuple(data.values())
                                elif isinstance(data, tuple):
                                    data_tuple = data
                                else:
                                    data_tuple = (data,)
                                patient_data_blood.append(data_tuple)
                                total_blood_files += 1
                            
                            # For vitals reports.
                            elif "VITALS" in first_page_text:
                                print("This is an vitals report from our reportingbot.")
                                # patient_data_vitals = extract_data_from_bot_vitals_file(first_page_text)
                                # print(patient_data_vitals)
                                # if isinstance(patient_data_vitals, dict):  # If it's a dictionary, convert it into a tuple or list
                                #     patient_data_vitals = tuple(patient_data_vitals.values())
                                #     print(f"Data after making it a tuple : {patient_data_vitals}")
                                # # else:
                                # #     data = patient_data_vitals
                                # total_vitals_files += 1
                                data = extract_data_from_bot_vitals_file(first_page_text)
                                if isinstance(data, dict):
                                    data_tuple = tuple(data.values())
                                elif isinstance(data, tuple):
                                    data_tuple = data
                                else:
                                    data_tuple = (data,)
                                patient_data_vitals.append(data_tuple)
                                total_vitals_files += 1

                            # For the xray report from stradus.
                            elif "Referrer Dr" in first_page_text and "Time" in first_page_text:
                                print("This is an xray report from our Stradus.")
                                # patient_data_xray = extract_data_from_the_stradus_xray_file(first_page_text)
                                # print(patient_data_xray)
                                # if isinstance(patient_data_xray, dict):  # If it's a dictionary, convert it into a tuple or list
                                #     patient_data_xray = tuple(patient_data_xray.values())
                                #     print(f"Data after making it a tuple : {patient_data_xray}")
                                # # else:
                                # #     data = patient_data_xray
                                # total_xray_files += 1
                                data = extract_data_from_the_stradus_xray_file(first_page_text)
                                if isinstance(data, dict):
                                    data_tuple = tuple(data.values())
                                elif isinstance(data, tuple):
                                    data_tuple = data
                                else:
                                    data_tuple = (data,)
                                patient_data_xray.append(data_tuple)
                                total_xray_files += 1
                            
                            # For the xray report from reporting bot.
                            elif "X-RAY" in first_page_text:
                                print("This is an xray report from our Reporting Bot.")
                                # patient_data_xray = extract_data_from_the_bot_xray_file(first_page_text)
                                # print(patient_data_xray)
                                # if isinstance(patient_data_xray, dict):  # If it's a dictionary, convert it into a tuple or list
                                #     patient_data_xray = tuple(patient_data_xray.values())
                                #     print(f"Data after making it a tuple : {patient_data_xray}")
                                # # else:
                                # #     data = patient_data_xray
                                # total_xray_files += 1
                                data = extract_data_from_the_bot_xray_file(first_page_text)
                                if isinstance(data, dict):
                                    data_tuple = tuple(data.values())
                                elif isinstance(data, tuple):
                                    data_tuple = data
                                else:
                                    data_tuple = (data,)
                                patient_data_xray.append(data_tuple)
                                total_xray_files += 1

                            # For Audiometry reports.
                            elif "left ear" in first_page_text:
                                print("This is an audio report from our Reporting Bot.")
                                # patient_data_audio = extract_data_from_bot_audio_file(first_page_text)
                                # print(patient_data_audio)
                                # if isinstance(patient_data_audio, dict):  # If it's a dictionary, convert it into a tuple or list
                                #     patient_data_audio = tuple(patient_data_audio.values())
                                #     print(f"Data after making it a tuple : {patient_data_audio}")
                                # # else:
                                # #     data = patient_data_audio
                                # total_audio_files += 1
                                data = extract_data_from_bot_audio_file(first_page_text)
                                if isinstance(data, dict):
                                    data_tuple = tuple(data.values())
                                elif isinstance(data, tuple):
                                    data_tuple = data
                                else:
                                    data_tuple = (data,)
                                patient_data_audio.append(data_tuple)
                                total_audio_files += 1

                            #ECG-REPORTINGBOT
                            elif "ECG" in first_page_text:
                                try:
                                    patient_id = str(first_page_text).split('Patient ID:')[1].split('Age:')[0].strip()
                                    patient_name = str(first_page_text).split("Name:")[1].split("Patient ID:")[0].strip()
                                    age = str(first_page_text).split("Age:")[1].split('Gender:')[0].strip()
                                    gender = str(first_page_text).split("Gender:")[1].split("Test Date:")[0].strip()
                                    test_date = str(first_page_text).split("Test Date:")[1].split('Report Date:')[0].strip()
                                    report_date = str(first_page_text).split("Report Date:")[1].split('ECG')[0].strip()

                                    # Extract heart rate using regex
                                    heart_rate = ""
                                    import re
                                    hr_match = re.search(r"Heart rate is\s*(\d+)\s*BPM", first_page_text, re.IGNORECASE)
                                    if hr_match:
                                        heart_rate = hr_match.group(1).strip()

                                    # Extract observations (line after heart rate)
                                    findings = ""
                                    observation_lines = first_page_text.split("Observation:")[1].strip().splitlines()
                                    cleaned_lines = [line.strip() for line in observation_lines if line.strip()]
                                    # Exclude the heart rate line and take the rest
                                    findings_lines = [line for line in cleaned_lines if "heart rate" not in line.lower()]
                                    if findings_lines:
                                        findings = findings_lines[0]  # First actual observation

                                    print(patient_id, patient_name, age, gender, test_date, report_date, heart_rate, findings)
                                    patient_data_ecg1.append((
                                        patient_id,
                                        patient_name,
                                        age,
                                        gender,
                                        test_date,
                                        report_date,
                                        heart_rate,
                                        remove_illegal_characters(findings)
                                    ))
                                    total_ecg_files1 += 1
                                except Exception as e:
                                    print(f"❌ ECG Extraction Error for {pdf_file.name}: {e}")
                                    error_count += 1
                                    error_file_path = error_dir / pdf_file.name
                                    shutil.copy2(pdf_file, error_file_path)


                            # If it is a Optometry report.
                            elif "OPTOMETRY" in first_page_text:
                                # patient_data_opto = extract_data_from_bot_opto_file(first_page_text)
                                # # If extract_data_from_bot_opto_file returns a dictionary, you may want to ensure it gets converted into a tuple or list for appending
                                # if isinstance(patient_data_opto, dict):  # If it's a dictionary, convert it into a tuple or list
                                #     data = tuple(patient_data_opto.values())
                                #     print(f"Data after making it a tuple : {data}")
                                # else:
                                #     data = patient_data_opto
                                # # patient_data_opto.append((patient_id, patient_name, age, gender, test_date, report_date, heart_rate, remove_illegal_characters(findings)))
                                # total_opto_files += 1
                                data = extract_data_from_bot_opto_file(first_page_text)
                                if isinstance(data, dict):
                                    data_tuple = tuple(data.values())
                                elif isinstance(data, tuple):
                                    data_tuple = data
                                else:
                                    data_tuple = (data,)
                                patient_data_opto.append(data_tuple)
                                total_opto_files += 1

                            else:
                                print(f"This is an Others File.")

                        except IndexError as e:
                            error_count += 1
                            error_file_path = error_dir / pdf_file.name
                            shutil.copy2(pdf_file, error_file_path)
                            print(f"Error processing file {pdf_file}: Invalid PDF Format")

            # ecg that we used to have on our drive (Not anymore needed.)
            if total_ecg_files > 0:
                workbook_ecg = openpyxl.Workbook()
                sheet_ecg = workbook_ecg.active

                sheet_ecg['A1'] = 'patient_id'
                sheet_ecg['B1'] = 'name'
                sheet_ecg['C1'] = 'age'
                sheet_ecg['D1'] = 'gender'
                sheet_ecg['E1'] = 'heart_rate'
                sheet_ecg['F1'] = 'report_time'
                sheet_ecg['G1'] = 'report_date'
                sheet_ecg['H1'] = 'R_II'
                sheet_ecg['I1'] = 'RR'
                sheet_ecg['J1'] = 'PR'
                sheet_ecg['K1'] = 'QRS'
                sheet_ecg['L1'] = 'QT'
                sheet_ecg['M1'] = 'QTc'
                sheet_ecg['N1'] = 'QT_QTc'

                for row, data in enumerate(patient_data_ecg, start=2):
                    sheet_ecg.append(data)

                excel_file_path_ecg = os.path.join(output_dir, "patient_data_ecg.xlsx")
                workbook_ecg.save(excel_file_path_ecg)

            # ecg that we are currently having in our reporting bot.
            if total_ecg_files1 > 0:
                workbook_ecg = openpyxl.Workbook()
                sheet_ecg = workbook_ecg.active

                sheet_ecg['A1'] = 'patient_id'
                sheet_ecg['B1'] = 'name'
                sheet_ecg['C1'] = 'age'
                sheet_ecg['D1'] = 'gender'
                sheet_ecg['E1'] = 'test_date'
                sheet_ecg['F1'] = 'report_date'
                sheet_ecg['G1'] = 'heart_rate'
                sheet_ecg['H1'] = 'findings'

                for row, data in enumerate(patient_data_ecg1, start=2):
                    sheet_ecg.append(data)

                excel_file_path_ecg = os.path.join(output_dir, "patient_data_ecg.xlsx")
                workbook_ecg.save(excel_file_path_ecg)

            # pft report data.
            if total_pft_files > 0:
                workbook_pft = openpyxl.Workbook()
                sheet_pft = workbook_pft.active

                sheet_pft['A1'] = 'patient_id'
                sheet_pft['B1'] = 'name'
                sheet_pft['C1'] = 'age'
                sheet_pft['D1'] = 'gender'
                sheet_pft['E1'] = 'height'
                sheet_pft['F1'] = 'weight'
                sheet_pft['G1'] = 'date'
                sheet_pft['H1'] = 'observation'

                for row, data in enumerate(patient_data_pft, start=2):
                    sheet_pft.append(data)

                excel_file_path_pft = os.path.join(output_dir, "patient_data_pft.xlsx")
                workbook_pft.save(excel_file_path_pft)

            # the optometry worksheet.
            if total_opto_files > 0:
                workbook_opto = openpyxl.Workbook()
                sheet_opto = workbook_opto.active

                sheet_opto['A1'] = 'patient_id'
                sheet_opto['B1'] = 'patient_name'
                sheet_opto['C1'] = 'patient_age'
                sheet_opto['D1'] = 'gender'
                sheet_opto['E1'] = 'test_date'
                sheet_opto['F1'] = 'report_date'
                sheet_opto['G1'] = 'far_vision_right'
                sheet_opto['H1'] = 'near_vision_right'
                sheet_opto['I1'] = 'distance_vision_right'
                sheet_opto['J1'] = 'reading_vision_right'
                sheet_opto['K1'] = 'spherical_right'
                sheet_opto['L1'] = 'cylindrical_right'
                sheet_opto['M1'] = 'axis_right'
                sheet_opto['N1'] = 'add_right'
                sheet_opto['O1'] = 'far_vision_left'
                sheet_opto['P1'] = 'near_vision_left'
                sheet_opto['Q1'] = 'distance_vision_right'
                sheet_opto['R1'] = 'reading_vision_left'
                sheet_opto['S1'] = 'spherical_left'
                sheet_opto['T1'] = 'cylindrical_left'
                sheet_opto['U1'] = 'axis_left'
                sheet_opto['V1'] = 'add_left'
                sheet_opto['W1'] = 'colour_blindness'

                for row, data in enumerate(patient_data_opto, start=2):
                    sheet_opto.append(data)

                excel_file_path_opto = os.path.join(output_dir, "OptometryPatientDetails.xlsx")
                workbook_opto.save(excel_file_path_opto)

            # the Pathology worksheet.
            if total_blood_files > 0:
                workbook_blood = openpyxl.Workbook()
                sheet_blood = workbook_blood.active

                sheet_blood['A1'] = 'patient_id'
                sheet_blood['B1'] = 'patient_name'
                sheet_blood['C1'] = 'patient_age'
                sheet_blood['D1'] = 'gender'
                sheet_blood['E1'] = 'test_date'
                sheet_blood['F1'] = 'report_date'
                sheet_blood['G1'] = 'haemoglobin'
                sheet_blood['H1'] = 'rbc_count'
                sheet_blood['I1'] = 'rbc_pcv'
                sheet_blood['J1'] = 'rbc_mcv'
                sheet_blood['K1'] = 'rbc_mch'
                sheet_blood['L1'] = 'rbc_mchc'
                sheet_blood['M1'] = 'rbc_rdw_cv'
                sheet_blood['N1'] = 'rbc_rdw_sd'
                sheet_blood['O1'] = 'wbc_tlc'
                sheet_blood['P1'] = 'dlc_neutrophils'
                sheet_blood['Q1'] = 'dlc_lymphocytes'
                sheet_blood['R1'] = 'dlc_monocytes'
                sheet_blood['S1'] = 'dlc_eosinophils'
                sheet_blood['T1'] = 'dlc_basophils'
                sheet_blood['U1'] = 'alc_neutrophils'
                sheet_blood['V1'] = 'alc_lymphocytes'
                sheet_blood['W1'] = 'alc_monocytes'
                sheet_blood['X1'] = 'alc_eosinophils'
                sheet_blood['Y1'] = 'alc_basophils'
                sheet_blood['Z1'] = 'platelet_count'
                sheet_blood['Y1'] = 'mean_platelet_volume'
                sheet_blood['Z1'] = 'pct'

                for row, data in enumerate(patient_data_blood, start=2):
                    sheet_blood.append(data)

                excel_file_path_blood = os.path.join(output_dir, "PathologyPatientDetails.xlsx")
                workbook_blood.save(excel_file_path_blood)

            # the audiometry worksheet.
            if total_audio_files > 0:
                workbook_audio = openpyxl.Workbook()
                sheet_audio = workbook_audio.active

                sheet_audio['A1'] = 'patient_id'
                sheet_audio['B1'] = 'patient_name'
                sheet_audio['C1'] = 'patient_age'
                sheet_audio['D1'] = 'gender'
                sheet_audio['E1'] = 'test_date'
                sheet_audio['F1'] = 'report_date'
                sheet_audio['G1'] = 'left_ear_finding'
                sheet_audio['H1'] = 'right_ear_finding'

                for row, data in enumerate(patient_data_audio, start=2):
                    sheet_audio.append(data)

                excel_file_path_audio = os.path.join(output_dir, "AudiometryPatientDetails.xlsx")
                workbook_audio.save(excel_file_path_audio)


            # the vitals worksheet.
            if total_vitals_files > 0:
                workbook_vitals = openpyxl.Workbook()
                sheet_vitals = workbook_vitals.active

                sheet_vitals['A1'] = 'patient_id'
                sheet_vitals['B1'] = 'patient_name'
                sheet_vitals['C1'] = 'patient_age'
                sheet_vitals['D1'] = 'gender'
                sheet_vitals['E1'] = 'test_date'
                sheet_vitals['F1'] = 'report_date'
                sheet_vitals['G1'] = 'height'
                sheet_vitals['H1'] = 'weight'
                sheet_vitals['I1'] = 'bp'
                sheet_vitals['J1'] = 'Pulse'

                for row, data in enumerate(patient_data_vitals, start=2):
                    sheet_vitals.append(data)

                excel_file_path_vitals = os.path.join(output_dir, "VitalsPatientDetails.xlsx")
                workbook_vitals.save(excel_file_path_vitals)


            green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # Green
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")    # Red

            if total_xray_files > 0:
                workbook_xray = openpyxl.Workbook()
                sheet_xray = workbook_xray.active
                sheet_xray.title = "X-Ray Reports"
                sheet_xray.append(['patient_id', 'name', 'age', 'gender', 'test_date', 'report_date', 'findings'])

                for data in patient_data_xray:
                    sheet_xray.append(data)

                # Define the reference phrase in lowercase
                # Define the reference phrases in lowercase
                target_phrases = ("No significant abnormality detected.", "No significant abnormality seen.")

                for row in range(2, len(patient_data_xray) + 2):
                    cell = sheet_xray.cell(row=row, column=7)
                    original_value = str(cell.value).strip() if cell.value else ""
                    normalized = original_value.strip()

                    # Check if the finding starts with any of the expected phrases
                    if normalized.startswith(target_phrases):
                        print(f"🟢 Normal Finding (Row {row - 1}): {original_value}")
                        cell.fill = green_fill
                    else:
                        if original_value:
                            print(f"🔴 Abnormal Finding (Row {row - 1}): {original_value}")
                        cell.fill = red_fill

                excel_file_path_xray = os.path.join(output_dir, "XrayPatientData.xlsx")
                workbook_xray.save(excel_file_path_xray)
                print(f"✅ Excel saved at: {excel_file_path_xray}")

            message = f"Total {total_ecg_files1} ECG and {total_pft_files} PFT and {total_xray_files} XRAY data and {total_opto_files} Opto and {total_vitals_files} Vitals data and {total_audio_files} Audio and {total_blood_files} Patho data files have been extracted and saved successfully.\n\n"
            message += f"ECG Output File: {excel_file_path_ecg}\n\nPFT Output File: {excel_file_path_pft}\n\nXRAY Output File: {excel_file_path_xray}"
            messagebox.showinfo("Patient Data Extractor", message)

        else:
            messagebox.showwarning("Output Folder Not Selected", "Output folder not selected.")
    else:
        messagebox.showwarning("Input Folder Not Selected", "Input folder not selected.")

# This is another function which reads the MERGED files one by one, and then tells for a patient, which files are present respectively, along with the data. - HIMANSHU.
def generate_excel_for_merged_files():
    # Call select_folders to get both input and output folder paths
    input_folder_path, output_folder_path = select_folders()
    if not input_folder_path or not output_folder_path:
        return
    # This will give me a list of only pdf files as the glob will only give me these, and then it will store them each in the list in the form of Path Object.
    pdf_files = list(Path(input_folder_path).glob("*.pdf"))
    print(pdf_files)
    # Getting the no. of files that i've processed.
    num_files_processed = len(pdf_files)
    # Later, i will convert the path object in the binary format so that i can read it using our reader and manipulate my data accordingly.

    # Collecting keys from the file names (extracted from the first part of the filename before the underscore)
    # I will use these in case if in any file the id is missing for that unique patient.
    keys = set()
    naming_errors = {}
    exception_files = {}
    incomplete_data = {}
    duplicate_file = {}
    id_mismatch = {}
    # Set to store modalities encountered for each unique file_id
    unique_file_id_set = set()

    # Making the modalities a set to store all the modalities for a particular id/key.
    modalities = set()

    # Initializing a empty dictionary to just store the patient details every time any file is processed.
    patient_details = create_patient_details()

    # Logic for merged files
    print(f"Input Folder: {input_folder_path}")
    print(f"Output Folder: {output_folder_path}")

    # Looping through the list of PDF files if those are already merged.
    for pdf_file in pdf_files:
        # This will extract the unique key from the file names.
        try:
            original_filename = str(pdf_file).split("\\")[-1]
            file_id = original_filename.split("_")[0].lower()
            if "." in file_id:
                naming_errors[str(file)] = original_filename
                print(f"File {pdf_file} has incorrect naming format. Storing naming error: {original_filename}")
                # Skipping to the next file in the loop, even if there is any naming error also, this will make sure that operations team do thier work properly.
                continue
            else:
                if file_id in keys:
                    # If file_id is already in the keys set, add it to the duplicate_file dictionary
                    duplicate_file[file_id] = original_filename
                    print(f"Duplicate file id : {file_id} found in File {pdf_file}, Skipping this file.")
                    # Skipping to the next file in the loop
                    continue
                else:
                    # Otherwise, adding the file_id to the keys set
                    keys.add(file_id)
        except IndexError:
            original_filename = str(pdf_file).split("\\")[-1]
            naming_errors[str(pdf_file)] = original_filename
            print(f"File {pdf_file} has incorrect naming format. Storing naming error: {original_filename}")
            # Skipping to the next file in the loop, even if there is any naming error also, this will make sure that operations team do thier work properly.
            continue
        print("Keys extracted from file names:", keys)
        print("Naming errors:", naming_errors)

        try:
            # Opening each PDF file in binary mode
            with open(pdf_file, 'rb') as file:
                # Creating a PdfReader object
                pdf_reader = PyPDF2.PdfReader(file)
                # Looping through each page in the PDF and save them as individual PDF files
                for page_number in range(len(pdf_reader.pages)):
                    # Extract text data from the page to determine the modality
                    page_text = pdf_reader.pages[page_number].extract_text()

                    # Log the page text for debugging, this will print every page.
                    # print_page_text_for_logging(page_text)

                    # This function i've created will check all the conditions and based on that give us the required details.
                    patient_details, modalities = extract_data_based_on_modality(page_text, patient_details, modalities)

                    for modality in modalities:
                        if modality not in unique_file_id_set:
                            # If this modality has not been processed for this file_id
                            print(f"Processing modality {modality} for file_id {file_id}")
                            # Add the modality to the unique set for this file_id
                            unique_file_id_set.add(modality)
                            # Call the modality based excel function to get the workbook for that modality
                            # Here the modality is storing the workbook name.
                            # i know that there isn't any workbook for Others so making this condition also worked.
                            if modality == 'OTHERS':
                                continue
                            # I've got the respective_patient_data, to store the data in proper way, might be useful.
                            modality, serial_no, respective_patient_data = modality_based_excel_workbook(modality)
                            # serial_no = modality.max_row + 1
                            # Creating the row data to append the data.
                            row_data = [serial_no] + list(patient_details.values())
                            # Make the workbook active to perform our work.
                            ws = modality.active
                            ws.append(row_data)
                            print(f"Added data for {modality} to row {serial_no}")
                    
                    # Clearing the modalities set when work is done.
                    modalities.clear()
        except Exception as e:
                print(f"Error processing {input_folder_path}: {str(e)}")
                exception_files[str(input_folder_path)] = str(e)
                continue  # Skip this file and continue with the next

    # Saving workbooks for each modality in the output directory
    modality_workbooks = {
        'XRAY': 'Xray_Test_Details.xlsx',
        'PFT': 'Pft_Test_Details.xlsx',
        'OPTOMETRY': 'Optometry_Test_Details.xlsx',
        'AUDIOMETRY': 'Audiometry_Test_Details.xlsx',
        'ECG': 'Ecg_Test_Details.xlsx',
        'VITALS': 'Vitals_Test_Details.xlsx',
        'PATHOLOGY': 'Pathology_Test_Details.xlsx'
    }

    # this is having some errors i need to fix them
    # Iterate over modality workbooks and save them
    for modality, filename in modality_workbooks.items():
        wb, _, _ = modality_based_excel_workbook(modality)
        if wb:
            output_file_path = Path(output_folder_path) / filename
            wb.save(output_file_path)
            print(f"Saved {modality} data to {output_file_path}")
    print(f"All the data extraction is completed and the errors are handled separately, and the data saved successfully to {output_file_path}.")
    
    # Display the completion message
    messagebox.showinfo("Process Completed", 
                        f"Total {num_files_processed} files were Processed.\n\nThe Excel file has been generated and saved to the selected Output Directory : \n{output_file_path}\n\n\nThank you for using OTHM !")
    
    # After processing files, calling the function to handle all errors
    handle_all_errors(naming_errors, duplicate_file, id_mismatch, incomplete_data, exception_files, output_folder_path, input_folder_path)


# This is the sample code to make a separate window to ask questions regarding which option our user wants to chose , 
# I'll use this afterwards.
# def count_of_tests_for_individual_patient():
#     # Create a simple custom window to ask the user to select an option
#     def on_select(option):
#         nonlocal selected_option
#         selected_option = option
#         window.destroy()  # Close the window once an option is selected

#     selected_option = None  # Variable to store the selected option

#     # Create a new Tkinter window for option selection
#     window = tk.Tk()
#     window.title("Choose Option")
#     window.geometry("300x150")

#     # Add two buttons for options
#     btn_individual = tk.Button(window, text="Count for Individual Files", command=lambda: on_select('individual'))
#     btn_individual.pack(pady=10)
    
#     btn_merged = tk.Button(window, text="Count for Merged Files", command=lambda: on_select('merged'))
#     btn_merged.pack(pady=10)

#     # Start the window's event loop
#     window.mainloop()

#     # If no option was selected, show a warning and exit
#     if selected_option is None:
#         messagebox.showwarning("No Option Chosen", "You must select an option!")
#         return

#     # Logic based on the selected option
#     if selected_option == 'individual':
#         # Logic for individual files
#         pdf_folder_path = filedialog.askdirectory(title="Select Individual Files Folder", mustexist=True)
#         if not pdf_folder_path:
#             print("Individual files folder not selected.")
#             return
#         # Add your logic for individual files here
        
#     elif selected_option == 'merged':
#         # Logic for merged files
#         pdf_folder_path = filedialog.askdirectory(title="Select Merged PDF Folder", mustexist=True)
#         if not pdf_folder_path:
#             print("Merged PDF folder not selected.")
#             return
#         # Add your logic for merged files here
# This function is used to check the data wrt to a particular excel, that if the data is matching correctly or not. -HIMANSHU.
def normalize_date(date_str):
    """
    Normalize dates in multiple formats to YYYY-MM-DD.
    """
    if not date_str:
        return None
    formats = ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d",
               "%d-%m-%y", "%Y.%m.%d", "%d %b %Y", "%d %B %Y"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime("%Y-%m-%d")
        except:
            continue
    return date_str.strip()

# Helper functions for data cleaning and comparison
def clean_age(age_str):
    """Clean age field by removing non-digit characters and leading zeros"""
    if not age_str:
        return ""
    
    # Remove non-digit characters
    digits = ''.join(filter(str.isdigit, str(age_str)))
    
    # Remove leading zeros and return
    if digits:
        # Convert to integer to remove leading zeros, then back to string
        return str(int(digits))
    return ""

def clean_gender(gender_str):
    """Clean gender field by standardizing format"""
    if not gender_str:
        return ""
    gender_str = str(gender_str).strip().lower()
    if gender_str in ['m', 'male']:
        return 'male'
    elif gender_str in ['f', 'female']:
        return 'female'
    return gender_str

def clean_name(name_str):
    """Clean name field by removing extra spaces, converting to lowercase, and removing common prefixes"""
    if not name_str:
        return ""
    
    # Convert to lowercase and remove extra spaces
    name_str = ' '.join(str(name_str).strip().lower().split())
    
    # Remove common prefixes
    prefixes = ['mr', 'mrs', 'miss', 'dr', 'md', 'prof']
    name_parts = name_str.split()
    if name_parts and name_parts[0] in prefixes:
        name_str = ' '.join(name_parts[1:])
    
    return name_str

def clean_id(id_str):
    """Clean ID field by converting to lowercase and stripping"""
    if not id_str:
        return ""
    return str(id_str).strip().lower()

def compare_fields(pdf_value, excel_value, field_name, problem_list, modality):
    """Compare two fields and add to problem list if they don't match"""
    if pdf_value != excel_value:
        problem_list.append(f"{modality}: {field_name} mismatch (PDF: '{pdf_value}' vs Excel: '{excel_value}')")
        return False
    return True

# This function is used to check the data wrt to a particular excel, that if the data is matching correctly or not. -HIMANSHU.
def check_pdf_files():
    # Asking for the I/P Directory ( files that need to be checked with the excel.)
    pdf_folder_path = filedialog.askdirectory(title="Select Merged PDF Folder", mustexist=True)
    if not pdf_folder_path:
        print("Merged PDF folder not selected.")
        return

    # Asking for the excel directory (the excel which is used for comparison.)
    excel_file_path = filedialog.askopenfilename(title="Select Excel Sheet", filetypes=[("Excel Files", "*.xlsx;*.xls")])
    if not excel_file_path:
        print("Excel sheet not selected.")
        return

    # Asking for the O/P Directory (here the compared excel will come.)
    output_directory = filedialog.askdirectory(title="Select Output Directory")
    if not output_directory:
        print("Output directory not selected.")
        return

    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ["patient_id", "patient_name", "age", "gender", "date", "ECG_GRAPH/ECG_REPORT", "XRAY_REPORT",
               "XRAY_IMAGE", "PFT", "AUDIOMETRY", "OPTOMETRY", "VITALS", "PROBLEM"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    comparison_df = pd.read_excel(excel_file_path)

    for _, excel_row in comparison_df.iterrows():
        pdf_id_prefix = str(excel_row['patient_id']).lower()

        pdf_files = [file.lower() for file in os.listdir(pdf_folder_path) if file.split("_")[0].lower() == pdf_id_prefix]

        # Initialize modality matching list outside the PDF page loop
        modality_match_list = []
        problem_list = []
        
        if not pdf_files:
            # No matching PDF file found for patient ID
            problem_list.append("Pdf file is missing")
            modality_match_list = ["No"] * 7

            # Use clean_age function to handle age normalization
            age_value = clean_age(str(excel_row.get("age", "")))

            # Write the results to the worksheet
            row_data = [
                           str(excel_row["patient_id"]).lower(),
                           str(excel_row["patient_name"]).split(" ")[0].lower(),
                           age_value,
                           str(excel_row["gender"]).strip(),
                           str(excel_row["date"]).strip(),
                       ] + modality_match_list + [', '.join(problem_list)]
            ws.append(row_data)

            current_row = ws.max_row

            yes_columns = [5, 6, 7, 8, 9, 10, 11, 12]  # columns E to L are modality columns

            # Apply fill color to cells based on "Yes" or "No"
            for col_num in range(5, 13):  # Columns E to L
                cell = ws.cell(row=current_row, column=col_num)
                if cell.value == "Yes":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00",
                                            fill_type="solid")  # Green color
                elif cell.value == "No":
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                            fill_type="solid")  # Red color
                    

        # To check whether pdf files are there or not.
        if pdf_files:
            pdf_file = pdf_files[0]
            pdf_path = os.path.join(pdf_folder_path, pdf_file)

            # Use clean_age function to handle age normalization
            age_value = clean_age(str(excel_row.get("age", "")))

            # Extract patient data from the Excel row
            patient_data_excel = {
                "patient_id": clean_id(excel_row["patient_id"]),
                "patient_name": clean_name(excel_row["patient_name"]),
                "age": age_value,  # Already cleaned with clean_age
                "gender": clean_gender(excel_row["gender"]),
                "date": normalize_date(str(excel_row["date"]).split(" ")[0])
            }
            print(patient_data_excel)

            pdf_reader = None  # Initialize PDF reader variable
            
            # Main Logic for comparison.
            try:
                # Open the PDF file once for all modalities
                try:
                    pdf_reader = PdfReader(open(pdf_path, "rb"))
                except Exception as e:
                    print(f"Error opening PDF file {pdf_file}: {str(e)}")
                    problem_list.append("Error opening PDF file")
                    modality_match_list = ["No"] * 7
                    # Write error row data
                    row_data = [
                        patient_data_excel["patient_id"],
                        patient_data_excel["patient_name"],
                        patient_data_excel["age"],
                        patient_data_excel["gender"],
                        patient_data_excel["date"]
                    ] + modality_match_list + [', '.join(problem_list)]
                    ws.append(row_data)
                    continue

                # Initialize modality matching for each modality
                modality_matches = {
                    "ECG_GRAPH/ECG_REPORT": False,
                    "XRAY_REPORT": False,
                    "XRAY_IMAGE": False,
                    "PFT": False,
                    "AUDIOMETRY": False,
                    "OPTOMETRY": False,
                    "VITALS": False
                }

                # Extract all text from PDF for better pattern matching
                full_text = ""
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    full_text += page.extract_text() + "\n"

                # Check for ECG - improved detection
                if "ECG" in full_text or "electrocardiogram" in full_text.lower():
                    try:
                        # Try to extract ECG data using multiple patterns
                        ecg_patterns = [
                            r"Patient ID:\s*([^\n]+)\s*Patient Name:\s*([^\n]+)\s*Age:\s*([^\n]+)\s*Gender:\s*([^\n]+)\s*Test Date:\s*([^\n]+)\s*Report Date:\s*([^\n]+)",
                            r"Patient ID:\s*([^\n]+)\s*Patient Name:\s*([^\n]+)\s*Age:\s*([^\n]+)\s*Gender:\s*([^\n]+)\s*Report Date:\s*([^\n]+)",
                            r"Name:\s*([^\n]+)\s*Patient ID:\s*([^\n]+)\s*Age:\s*([^\n]+)\s*Gender:\s*([^\n]+)\s*Report date:\s*([^\n]+)"
                        ]
                        
                        for pattern in ecg_patterns:
                            match = re.search(pattern, full_text, re.IGNORECASE)
                            if match:
                                groups = match.groups()
                                if len(groups) >= 5:
                                    if len(groups) == 6:
                                        patient_id = clean_id(groups[0])
                                        patient_name = clean_name(groups[1])
                                        age = clean_age(groups[2])
                                        gender = clean_gender(groups[3])
                                        report_date = normalize_date(groups[5])  # Use Report Date
                                    else:
                                        patient_id = clean_id(groups[0])
                                        patient_name = clean_name(groups[1])
                                        age = clean_age(groups[2])
                                        gender = clean_gender(groups[3])
                                        report_date = normalize_date(groups[4])
                                    
                                    print("ECG details found:", patient_id, patient_name, age, gender, report_date)
                                    
                                    # Compare with Excel data
                                    excel_date = patient_data_excel["date"]
                                    pdf_date = normalize_date(report_date)
                                    
                                    # Compare with Excel data and track mismatches
                                    all_match = True
                                    all_match &= compare_fields(patient_id, patient_data_excel["patient_id"], "Patient ID", problem_list, "ECG")
                                    all_match &= compare_fields(patient_name, patient_data_excel["patient_name"], "Patient Name", problem_list, "ECG")
                                    all_match &= compare_fields(age, patient_data_excel["age"], "Age", problem_list, "ECG")
                                    all_match &= compare_fields(gender, patient_data_excel["gender"], "Gender", problem_list, "ECG")
                                    all_match &= compare_fields(pdf_date, excel_date, "Date", problem_list, "ECG")
                                    
                                    if all_match:
                                        modality_matches["ECG_GRAPH/ECG_REPORT"] = True
                                        break
                                break
                    except Exception as e:
                        print(f"Error processing ECG: {str(e)}")
                        problem_list.append(f"ECG: Error processing - {str(e)}")

                # Check for Audiometry - improved detection
                if "audiometry" in full_text.lower() or "left ear" in full_text.lower() or "right ear" in full_text.lower():
                    try:
                        # Try to extract Audiometry data using multiple patterns
                        audio_patterns = [
                            r"Patient ID:\s*([^\n]+)\s*Patient Name:\s*([^\n]+)\s*Age:\s*([^\n]+)\s*Gender:\s*([^\n]+)\s*Test Date:\s*([^\n]+)\s*Report Date:\s*([^\n]+)",
                            r"Patient ID:\s*([^\n]+)\s*Patient Name:\s*([^\n]+)\s*Age:\s*([^\n]+)\s*Gender:\s*([^\n]+)\s*Report Date:\s*([^\n]+)",
                            r"Name:\s*([^\n]+)\s*Patient ID:\s*([^\n]+)\s*Age:\s*([^\n]+)\s*Gender:\s*([^\n]+)\s*Report date:\s*([^\n]+)"
                        ]
                        
                        for pattern in audio_patterns:
                            match = re.search(pattern, full_text, re.IGNORECASE)
                            if match:
                                groups = match.groups()
                                if len(groups) >= 5:
                                    if len(groups) == 6:
                                        patient_id = clean_id(groups[0])
                                        patient_name = clean_name(groups[1])
                                        age = clean_age(groups[2])
                                        gender = clean_gender(groups[3])
                                        report_date = normalize_date(groups[5])  # Use Report Date
                                    else:
                                        patient_id = clean_id(groups[0])
                                        patient_name = clean_name(groups[1])
                                        age = clean_age(groups[2])
                                        gender = clean_gender(groups[3])
                                        report_date = normalize_date(groups[4])
                                    
                                    print("Audiometry details found:", patient_id, patient_name, age, gender, report_date)
                                    
                                    # Compare with Excel data
                                    excel_date = patient_data_excel["date"]
                                    pdf_date = normalize_date(report_date)
                                    
                                    # Compare with Excel data and track mismatches
                                    all_match = True
                                    all_match &= compare_fields(patient_id, patient_data_excel["patient_id"], "Patient ID", problem_list, "AUDIOMETRY")
                                    all_match &= compare_fields(patient_name, patient_data_excel["patient_name"], "Patient Name", problem_list, "AUDIOMETRY")
                                    all_match &= compare_fields(age, patient_data_excel["age"], "Age", problem_list, "AUDIOMETRY")
                                    all_match &= compare_fields(gender, patient_data_excel["gender"], "Gender", problem_list, "AUDIOMETRY")
                                    all_match &= compare_fields(pdf_date, excel_date, "Date", problem_list, "AUDIOMETRY")
                                    
                                    if all_match:
                                        modality_matches["AUDIOMETRY"] = True
                                        break
                                break
                    except Exception as e:
                        print(f"Error processing AUDIOMETRY: {str(e)}")
                        problem_list.append(f"AUDIOMETRY: Error processing - {str(e)}")

                # Process other modalities as before
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()

                    print("this is the start of page text.")
                    print("Page no. ", page_num)
                    print(page_text)
                    print("This is the end of page text.")

                    # checking for pft.
                    try:
                        print("inside the try block of pft.")
                        if not modality_matches["PFT"] and "RECORDERS & MEDICARE SYSTEMS" in page_text:
                            print("it confirms that it is a pft file.")
                            patient_name = clean_name(str(page_text).split("Patient: ")[1].split("Refd.By:")[0].split("\n")[0])
                            # Naming Issue required by team, making sure that there will be no space in the name.- Himanshu.
                            if " " in patient_name:
                                patient_name = patient_name.split(" ")[0]
                            patient_id = clean_id(str(page_text).split("ID     :")[1].split("Age")[0])
                            age = clean_age(str(page_text).split("Age    :")[1].split("Yrs")[0])
                            if "Smoker" in page_text:
                                gender = clean_gender(str(page_text).split("Gender   :")[1].split("Smoker")[0])
                            else:
                                gender = clean_gender(str(page_text).split("Gender   :")[1].split("Eth. Corr:")[0])
                            date = str(page_text).split("Date   :")[1][1:13].strip().lower()
                            if len(date) == 10:
                                report_date = date
                            else:
                                try:
                                    input_date = datetime.strptime(date, "%d-%b-%Y")
                                    report_date = input_date.strftime("%Y-%m-%d")
                                except:
                                    report_date = date

                            print('PFT', patient_id, patient_name, age, gender, report_date)

                            # Normalize dates before comparison
                            excel_date = patient_data_excel["date"]
                            pdf_date = normalize_date(report_date)

                            # Compare with Excel data and track mismatches
                            all_match = True
                            all_match &= compare_fields(patient_id, patient_data_excel["patient_id"], "Patient ID", problem_list, "PFT")
                            all_match &= compare_fields(patient_name, patient_data_excel["patient_name"], "Patient Name", problem_list, "PFT")
                            all_match &= compare_fields(age, patient_data_excel["age"], "Age", problem_list, "PFT")
                            all_match &= compare_fields(gender, patient_data_excel["gender"], "Gender", problem_list, "PFT")
                            all_match &= compare_fields(pdf_date, excel_date, "Date", problem_list, "PFT")
                            
                            if all_match:
                                modality_matches["PFT"] = True
                                    
                    except IndexError as ie:
                        print(f"IndexError: {str(ie)}. Skipping page processing.")
                        continue
                    except Exception as e:
                        print(f"Error processing PFT: {str(e)}")
                        continue

                    # Checking for opto.
                    try:
                        print("Inside the try block of optometry.")
                        if not modality_matches["OPTOMETRY"] and "OPTOMETRY REPORT" in page_text:
                            print("This is confirmed that this is a opto file.")
                            patient_name = clean_name(str(page_text).split("Name:")[1].split("Age:")[0])
                            patient_id = clean_id(str(page_text).split("Patient ID:")[1].split("Patient Name:")[0])
                            age = clean_age(str(page_text).split("Age:")[1].split("Gender")[0])
                            gender = clean_gender(str(page_text).split("Gender:")[1].split("Test")[0])
                            report_date = normalize_date(str(page_text).split("Report Date:")[1].split("OPTOMETRY")[0])

                            print("These are the opto patient details :")
                            print("Patient Id", patient_id)
                            print("Patient Name", patient_name)
                            print("Age", age)
                            print("Gender", gender)
                            print("Report Date", report_date)
                            
                            print('OPTOMETRY', patient_id, patient_name, age, gender, report_date)
                            # Normalize dates before comparison
                            excel_date = patient_data_excel["date"]
                            pdf_date = normalize_date(report_date)
                            
                            # Compare with Excel data and track mismatches
                            all_match = True
                            all_match &= compare_fields(patient_id, patient_data_excel["patient_id"], "Patient ID", problem_list, "OPTOMETRY")
                            all_match &= compare_fields(patient_name, patient_data_excel["patient_name"], "Patient Name", problem_list, "OPTOMETRY")
                            all_match &= compare_fields(age, patient_data_excel["age"], "Age", problem_list, "OPTOMETRY")
                            all_match &= compare_fields(gender, patient_data_excel["gender"], "Gender", problem_list, "OPTOMETRY")
                            all_match &= compare_fields(pdf_date, excel_date, "Date", problem_list, "OPTOMETRY")
                            
                            if all_match:
                                modality_matches["OPTOMETRY"] = True

                    except IndexError as ie:
                        print(f"IndexError: {str(ie)}. Skipping page processing.")
                        continue
                    except Exception as e:
                        print(f"Error processing OPTOMETRY: {str(e)}")
                        continue

                    # Checking for vitals.
                    try:
                        print("inside the try block of vitals.")
                        if not modality_matches["VITALS"] and "VITAL" in page_text:
                            print("it confirms that it is a vitals file.")
                            patient_id = clean_id(str(page_text).split("Patient ID:")[1].split("Patient Name:")[0])
                            patient_name = clean_name(str(page_text).split("Patient Name:")[1].split("Age")[0])
                            age = clean_age(str(page_text).split("Age:")[1].split("Gender")[0])
                            gender = clean_gender(str(page_text).split("Gender:")[1].split("Test")[0])
                            report_date = normalize_date(str(page_text).split("Report Date:")[1].split("VITALS")[0])
                            print('VITALS', patient_id, patient_name, age, gender, report_date)
                            # Normalize dates
                            excel_date = patient_data_excel["date"]
                            pdf_date = normalize_date(report_date)
                            
                            # Compare with Excel data and track mismatches
                            all_match = True
                            all_match &= compare_fields(patient_id, patient_data_excel["patient_id"], "Patient ID", problem_list, "VITALS")
                            all_match &= compare_fields(patient_name, patient_data_excel["patient_name"], "Patient Name", problem_list, "VITALS")
                            all_match &= compare_fields(age, patient_data_excel["age"], "Age", problem_list, "VITALS")
                            all_match &= compare_fields(gender, patient_data_excel["gender"], "Gender", problem_list, "VITALS")
                            all_match &= compare_fields(pdf_date, excel_date, "Date", problem_list, "VITALS")
                            
                            if all_match:
                                modality_matches["VITALS"] = True
                    except IndexError as ie:
                        print(f"IndexError: {str(ie)}. Skipping page processing.")
                        continue
                    except Exception as e:
                        print(f"Error processing VITALS: {str(e)}")
                        continue

                    # Checking for X-Ray (Reporting Bot.)
                    # Checking for X-Ray (Reporting Bot.)
                    try:
                        print("Inside the try block of XRAY.")
                        if not modality_matches["XRAY_REPORT"] and "X-RAY" in page_text:
                            print("Confirmed X-Ray file.")
                            
                            # Extract Patient Fields with more flexible parsing
                            patient_id = None
                            patient_name = None
                            age = None
                            gender = None
                            test_date = None
                            report_date = None
                            
                            # Try multiple patterns to extract patient information
                            patterns = [
                                r"Patient Name:\s*([^\n]+)\s*Patient ID:\s*([^\n]+)\s*Patient Age:\s*([^\n]+)\s*Patient Gender:\s*([^\n]+)\s*Test Date:\s*([^\n]+)\s*Report Date:\s*([^\n]+)",
                                r"Patient Name:\s*([^\n]+)\s*Patient ID:\s*([^\n]+)\s*Patient Age:\s*([^\n]+)\s*Patient Gender:\s*([^\n]+)",
                                r"Patient ID:\s*([^\n]+)\s*Patient Name:\s*([^\n]+)\s*Patient Age:\s*([^\n]+)\s*Patient Gender:\s*([^\n]+)"
                            ]
                            
                            for pattern in patterns:
                                match = re.search(pattern, page_text, re.IGNORECASE)
                                if match:
                                    groups = match.groups()
                                    if len(groups) >= 4:
                                        if "Patient Name:" in pattern:
                                            patient_name = clean_name(groups[0])
                                            patient_id = clean_id(groups[1])
                                            age = clean_age(groups[2])
                                            gender = clean_gender(groups[3])
                                            if len(groups) >= 6:
                                                test_date = normalize_date(groups[4])
                                                report_date = normalize_date(groups[5])
                                        else:  # Patient ID first pattern
                                            patient_id = clean_id(groups[0])
                                            patient_name = clean_name(groups[1])
                                            age = clean_age(groups[2])
                                            gender = clean_gender(groups[3])
                                    break
                            
                            # If regex didn't work, try the original method with more error handling
                            if not all([patient_id, patient_name, age, gender]):
                                try:
                                    patient_id = clean_id(str(page_text).split("Patient ID:")[1].split("Patient Age:")[0].strip())
                                except:
                                    pass
                                try:
                                    patient_name = clean_name(str(page_text).split("Patient Name:")[1].split("Patient ID:")[0].strip())
                                except:
                                    pass
                                try:
                                    age = clean_age(str(page_text).split("Patient Age:")[1].split("Patient Gender:")[0].strip().lstrip('0'))
                                except:
                                    pass
                                try:
                                    gender = clean_gender(str(page_text).split("Patient Gender:")[1].split("Test Date:")[0].strip())
                                except:
                                    pass
                                try:
                                    test_date = normalize_date(str(page_text).split("Test Date:")[1].split("Report Date:")[0].strip())
                                except:
                                    pass
                                try:
                                    report_date = normalize_date(str(page_text).split("Report Date:")[1].split("Referral Dr:")[0].strip())
                                except:
                                    pass

                            print('XRAY BOT:', patient_id, patient_name, age, gender, report_date or test_date)

                            # Normalize dates
                            excel_date = patient_data_excel["date"]
                            pdf_date = normalize_date(test_date) if test_date else normalize_date(report_date)

                            # Compare with Excel data
                            excel_data = patient_data_excel
                            
                            # Compare with Excel data and track mismatches
                            all_match = True
                            if patient_id:
                                all_match &= compare_fields(patient_id, excel_data["patient_id"], "Patient ID", problem_list, "XRAY")
                            else:
                                problem_list.append("XRAY: Patient ID not found")
                                
                            if patient_name:
                                all_match &= compare_fields(patient_name, excel_data["patient_name"], "Patient Name", problem_list, "XRAY")
                            else:
                                problem_list.append("XRAY: Patient Name not found")
                                
                            if age:
                                all_match &= compare_fields(age, excel_data["age"], "Age", problem_list, "XRAY")
                            else:
                                problem_list.append("XRAY: Age not found")
                                
                            if gender:
                                all_match &= compare_fields(gender, excel_data["gender"], "Gender", problem_list, "XRAY")
                            else:
                                problem_list.append("XRAY: Gender not found")
                                
                            if pdf_date:
                                all_match &= compare_fields(pdf_date, excel_date, "Date", problem_list, "XRAY")
                            else:
                                problem_list.append("XRAY: Date not found")

                            if all_match:
                                modality_matches["XRAY_REPORT"] = True

                    except Exception as e:
                        print(f"Error processing XRAY: {str(e)}")
                        import traceback
                        traceback.print_exc()
                    # Check for XRAY_IMAGE
                    try:
                        if not modality_matches["XRAY_IMAGE"] and "Page 2 of 2" in page_text:
                            if "Page 2 of 2" in page_text:
                                modality_matches["XRAY_IMAGE"] = True
                    except IndexError as ie:
                        print(f"IndexError: {str(ie)}. Skipping page processing.")
                        continue

                # Convert modality matches to list for Excel output
                modality_match_list = [
                    "Yes" if modality_matches["ECG_GRAPH/ECG_REPORT"] else "No",
                    "Yes" if modality_matches["XRAY_REPORT"] else "No", 
                    "Yes" if modality_matches["XRAY_IMAGE"] else "No",
                    "Yes" if modality_matches["PFT"] else "No",
                    "Yes" if modality_matches["AUDIOMETRY"] else "No",
                    "Yes" if modality_matches["OPTOMETRY"] else "No",
                    "Yes" if modality_matches["VITALS"] else "No"
                ]

                # Generate problem list for mismatches
                for modality, matched in modality_matches.items():
                    if not matched and modality != "XRAY_IMAGE":  # XRAY_IMAGE might not have patient data
                        if not any(modality in problem for problem in problem_list):
                            problem_list.append(f"{modality}: Not found")

                # Write the results to the worksheet
                row_data = [
                    patient_data_excel["patient_id"],
                    patient_data_excel["patient_name"],
                    patient_data_excel["age"],
                    patient_data_excel["gender"],
                    patient_data_excel["date"]
                ] + modality_match_list + [', '.join(problem_list)]
                ws.append(row_data)

                current_row = ws.max_row

                # Apply fill color to cells based on "Yes" or "No"
                for col_num in range(6, 13):  # Columns F to L (modality columns)
                    cell = ws.cell(row=current_row, column=col_num)
                    if cell.value == "Yes":
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00",
                                                fill_type="solid")  # Green color
                    elif cell.value == "No":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color

            except Exception as e:
                print(f"Error processing PDF file {pdf_file}: {str(e)}")
                # Close PDF reader if it was opened
                if pdf_reader and hasattr(pdf_reader, 'stream') and pdf_reader.stream:
                    try:
                        pdf_reader.stream.close()
                    except:
                        pass
                
                # Move file to error folder with proper error handling
                error_folder_path = os.path.join(output_directory, "error")
                os.makedirs(error_folder_path, exist_ok=True)
                
                try:
                    # Try to close any open file handles first
                    import gc
                    gc.collect()  # Force garbage collection
                    
                    # Use copy instead of move to avoid permission issues
                    import shutil
                    error_file_path = os.path.join(error_folder_path, pdf_file)
                    shutil.copy2(pdf_path, error_file_path)
                    print(f"Copied {pdf_file} to error folder")
                except Exception as move_error:
                    print(f"Could not move file to error folder: {str(move_error)}")
                    
                continue
        else:
            print(f"No matching PDF file found for patient ID: {pdf_id_prefix}")

    # Save the workbook to the output directory
    output_filename = "patient_data_comparison.xlsx"
    wb.save(os.path.join(output_directory, output_filename))
    print("Data comparison completed.")
    messagebox.showinfo("Process Completed", "Data comparison completed.")

def sanitize_filename(filename):
    # Remove invalid characters from the filename
    return re.sub(r'[\\/:*?"<>|]', '_', filename)

def split_patient_file():
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            pdf_dir = Path(input_directory)
            pdf_output_dir = Path(output_directory)
            pdf_output_dir.mkdir(parents=True, exist_ok=True)

            pdf_files = list(pdf_dir.glob("*.pdf"))

            if pdf_files:
                for input_pdf_path in pdf_files:
                    try:
                        # Open the merged PDF file
                        with open(input_pdf_path, 'rb') as pdf_file:
                            pdf_reader = PyPDF2.PdfReader(pdf_file)

                            # Create a subdirectory for the patient if it doesn't exist
                            patient_id = sanitize_filename(input_pdf_path.stem)
                            patient_dir = pdf_output_dir / patient_id
                            patient_dir.mkdir(parents=True, exist_ok=True)

                            # Loop through each page in the PDF and save them as individual PDF files
                            for page_number in range(len(pdf_reader.pages)):
                                pdf_writer = PyPDF2.PdfWriter()
                                pdf_writer.add_page(pdf_reader.pages[page_number])

                                # Extract text data from the page to determine the modality
                                page_text = pdf_reader.pages[page_number].extract_text()
                                modality = None
                                images = []
                                is_ecg = False
                                count = 0

                                # Adding logs so that i can identify the page. - himanshu.
                                print("Page text is below")
                                print(page_text)
                                print("End of page text.")

                                if "SMART REPORT" in page_text:
                                    print("This is a smart report.")
                                    modality = "Smart_Report"
                                elif "X-RAY" in page_text:
                                    print("This is a xray file.")
                                    modality = 'Xray_Report'
                                elif "RECORDERS & MEDICARE SYSTEMS" in page_text:
                                    print("This is a pft file.")
                                    modality = 'PFT'
                                elif "Page 2 of 2" in page_text:
                                    print("This is an xray image.")
                                    modality = 'Xray_Image'
                                elif "OPTOMETRY" in page_text:
                                    print("This is a optometry file.")
                                    modality = 'Optometry'
                                elif "left ear" in page_text:
                                    print("This is a audiometry file.")
                                    modality = 'Audiometry'
                                elif "ECG" in page_text:
                                    print("This is an ECG file.")
                                    modality = 'ECG'
                                elif is_ecg == True and count == 1:
                                    print("This is another others image.")
                                    modality = 'Other2'
                                    count += 1
                                elif is_ecg == True and count == 2:
                                    print("This is the 3rd others file.")
                                    modality = 'Others3'
                                elif page_text == '':
                                    print("This is a others image.")
                                    modality = 'Others'
                                    is_ecg = True
                                    count += 1
                                elif "VITALS" in page_text:
                                    print("This is a vitals file.")
                                    modality = 'Vitals'
                                elif "RBC Count" in page_text:
                                    print("This is a blood Report.")
                                    modality = 'BloodReport'
                                else:
                                    print("This is a dr. consultation  file.")
                                    modality = 'Dr.Consultation'

                                # I have to add a proper logic to separate the images separately.

                                if modality:
                                    output_file_path = patient_dir / f'{patient_id}_{modality}.pdf'
                                else:
                                    output_file_path = patient_dir / f'{patient_id}_page_{page_number + 1}.pdf'

                                # Save the individual page as a PDF file with the new name
                                with open(output_file_path, 'wb') as output_file:
                                    pdf_writer.write(output_file)

                            print(f"PDF files for patient {patient_id} split and renamed successfully.")
                    except Exception as e:
                        print(f"Error processing {input_pdf_path}: {str(e)}")
                        continue  # Skip this file and continue with the next

                print("PDF files processed.")
            else:
                print("No PDF files found in the input directory.")

        else:
            print("Output directory not selected.")
    else:
        print("Input directory not selected.")

# This is my another function which gives me the respective data i.e. excel sheet for each test.
# Ex. appending all the optometry patients list in the excel along with the report findings, and similarly others.
def generate_patient_report_excel():
    return 'Himanshu is working on it'

# This is my new function which will just give me the count or tell me whether what test's are done for a particular patient. - Himanshu.
def count_of_tests_for_individual_patient():
    # As of now , i am not able to use a simple dialogbox to make the user select option from my window directly.
    # I've also thought of modifying the Yes and NO option in the ' messagebox.askquestion() ', but i guess I am not able to change these labels in tkinter directly as of now.
    # If i get any other way to change these than i will use that part directly instead of creating a separate dialogbox.
    
    # This is the code that just opens the separate window for selecting the option.
    # Creating a new Tkinter window for option selection.
    optionWindow = tk.Toplevel(window)
    optionWindow.title("Select an Option")
    optionWindow.geometry("300x150")

    # Centering the optionWindow on the main window.
    window_width = window.winfo_width()
    window_height = window.winfo_height()
    window_x = window.winfo_x()
    window_y = window.winfo_y()
    
    # Calculating the position of the option window to be centered
    option_window_x = window_x + (window_width // 2) - 150  # 150 is half the width of option window (300x150)
    option_window_y = window_y + (window_height // 2) - 75  # 75 is half the height of option window
    
    optionWindow.geometry(f"300x150+{option_window_x}+{option_window_y}")

    def option_selected(option):
        nonlocal selected_option
        selected_option = option
        optionWindow.destroy()  # Close the option selection window
    
    selected_option = None  # Variable to store the selected option

    # Add buttons for options
    btn_individual = tk.Button(optionWindow, text="Count for Individual Files", command=lambda: option_selected(1), bg="gray", width=25)
    btn_individual.pack(pady=20)
    
    btn_merged = tk.Button(optionWindow, text="Count for Merged Files", command=lambda: option_selected(2), bg="gray", width=25)
    btn_merged.pack(pady=10)

    # Start the window's event loop
    optionWindow.grab_set()
    # optionWindow.mainloop()
    # Wait for the option window to be closed before continuing with the main program
    window.wait_window(optionWindow)

    # End of the separate selection window code.

    # Start of the option respective logic.

    # If no option was selected, show a warning and exit
    if selected_option is None:
        messagebox.showwarning("No Option Chosen", "You must select an option to continue with Excel Generation!")
        return
    
    # Call select_folders to get both input and output folder paths
    input_folder_path, output_folder_path = select_folders()
    if not input_folder_path or not output_folder_path:
        return

    # Now, this will give me a list of only pdf files as the glob will only give me these, and then it will store them each in the list in the form of Path Object.
    pdf_files = list(Path(input_folder_path).glob("*.pdf"))
    print(pdf_files)
    # Getting the no. of files that i've processed.
    num_files_processed = len(pdf_files)
    # Later, i will convert the path object in the binary format so that i can read it using our reader and manipulate my data accordingly.

    # Collecting keys from the file names (extracted from the first part of the filename before the underscore)
    # I will use these in case if in any file the id is missing for that unique patient.
    keys = set()
    naming_errors = {}
    exception_files = {}
    incomplete_data = {}
    duplicate_file = {}
    id_mismatch = {}

    # Making the modalities a set to store all the modalities for a particular id/key.
    modalities = set()

    # Initializing a empty dictionary to just store the patient details every time any file is processed.
    patient_details = create_patient_details()
    # patient_details= {'patient_id': None,'patient_name': None,'patient_age': None,'gender': None,'test_date': None,'report_date': None}


    # defining some unwanted phrases for later use.
    # i've defined in the conditional function, so i might not require here now.
    # unwanted_phrases = ["Referrer Dr.","left ear","RECORDERS & MEDICARE SYSTEMS","OPTOMETRY","ECG","VITALS","RBC Count","PDW *","PDW"]

    # Initializing the patient_data dictionary before looping through each and every file so that i can use it to fill the excel.
    patient_data = creating_or_emptying_the_patient_data_dictionary()
    
    # Excel Workbook Setup.
    wb = Workbook()
    ws = wb.active

    # As of now, i have total 15 headers only.
    headers = ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE', 
            'XRAY', 'ECG', 'AUDIOMETRY', 'OPTOMETRY', 'VITALS', 'PFT', 'PATHOLOGY', 'OTHERS']

    # Adding headers in excel
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Initializing the serial number for patient data.
    serial_no = 1


    # Logic based on the selected option
    if selected_option == 1:
        # Logic for individual files
        print(f"Input Folder: {input_folder_path}")
        print(f"Output Folder: {output_folder_path}")

        # Looping through the list of PDF files if those are not merged yet.
        for pdf_file in pdf_files:
            # This will extract the unique key from the file names.
            try:
                original_filename = str(pdf_file).split("\\")[-1]
                file_id = original_filename.split("_")[0].lower()
                if "." in file_id:
                    naming_errors[str(file)] = original_filename
                    print(f"File {pdf_file} has incorrect naming format. Storing naming error: {original_filename}")
                    # Skipping to the next file in the loop, even if there is any naming error also, this will make sure that operations team do thier work properly.
                    continue
                else:
                    if file_id in keys:
                        # setting the patient_details and modalities for the current file_id, which is already created.
                        patient_details = patient_data[file_id]["patient_details"]
                        modalities = patient_data[file_id]["modalities"]
                    else:
                        # Otherwise, adding the file_id to the keys set
                        keys.add(file_id)
                        # Creating a new patient details dictionary and an empty set for modalities
                        patient_data[file_id] = {
                            "patient_details": create_patient_details(),
                            "modalities": set()
                        }

                        # setting the patient_details and modalities for the current file_id, this will be unique.
                        patient_details = patient_data[file_id]["patient_details"]
                        modalities = patient_data[file_id]["modalities"]

            except IndexError:
                original_filename = str(pdf_file).split("\\")[-1]
                naming_errors[str(pdf_file)] = original_filename
                print(f"File {pdf_file} has incorrect naming format. Storing naming error: {original_filename}")
                # Skipping to the next file in the loop, even if there is any naming error also, this will make sure that operations team do thier work properly.
                continue
            print("Keys extracted from file names:", keys)
            print("Naming errors:", naming_errors)

            try:
                # Opening each PDF file in binary mode
                with open(pdf_file, 'rb') as file:
                    # Creating a PdfReader object
                    pdf_reader = PyPDF2.PdfReader(file)
                    # Looping through each page in the PDF and save them as individual PDF files
                    for page_number in range(len(pdf_reader.pages)):
                        # Extract text data from the page to determine the modality
                        page_text = pdf_reader.pages[page_number].extract_text()

                        # Log the page text for debugging, this will print every page.
                        # print_page_text_for_logging(page_text)

                        # This function i've created will check all the conditions and based on that give us the required details.
                        patient_details, modalities = extract_data_based_on_modality(page_text, patient_details, modalities)

                        # Checking if patient_id matches with the file_id
                        if patient_details['patient_id'] != None:
                            if patient_details['patient_id'] != file_id:
                                id_extracted = patient_details['patient_id']
                                id_mismatch[id_extracted] = original_filename
                                print(f"Id in File : {pdf_file} and in it's filename is not matching, Please Review this file.")

                        # Update the patient_data dictionary for this file_id
                        patient_data[file_id]["patient_details"] = patient_details
                        patient_data[file_id]["modalities"] = modalities

            except Exception as e:
                print(f"Error processing {input_folder_path}: {str(e)}")
                exception_files[str(input_folder_path)] = str(e)
                continue  # Skip this file and continue with the next
        
        for file_id in keys:
            # setting the patient_details and modalities for the current file_id, which is already created.
            patient_details = patient_data[file_id]["patient_details"]
            modalities = patient_data[file_id]["modalities"]
            # Now, checking that is there any "None" or empty value in the patient details, if yes , that means there is incomplete data in it.
            missing_keys = [key for key, value in patient_details.items() if value is None]
            if missing_keys:
                incomplete_data[file_id] = original_filename
                print(f"Incomplete Data found in file id : {file_id} in File {pdf_file}, Please Review this file.")

            # Now, I'll update the data in the patien_data dictionary so that i can simply use it to put it in the excel.
            # After processing the pages of the current PDF file, just before moving to the next file:

            # creating the unique patient data for each file so that i can use that instead of the globally available thing.
            each_person_patient_data = creating_or_emptying_the_patient_data_dictionary()
            # Updating patient_data from patient_details
            for key in patient_details:
                each_person_patient_data[key] = patient_details[key]
                # if patient_details[key] is not None:
                #     patient_data[key] = patient_details[key]

            # i have to update the original_filename here for each file id.
            # Checking the modalities set and update corresponding fields in patient_data
            for modality in modalities:
                if modality in patient_data:
                    each_person_patient_data[modality] = 'Present'

            # I'll further process these now, as of now , printing these for additional logs.
            print(f"Patient data for {file_id}: {each_person_patient_data}")
            
            # These things are not needed because these will be automatically handled by the python garbage collector.
            # Clearing the modalities set and patient_details dictionary for the next file
            # patient_data[file_id]["modalities"].clear()
            # # Resetting values to None
            # for key in patient_data[file_id]["patient_details"]:
            #     patient_data[file_id]["patient_details"][key] = None

            # print(f"(This is the confirmation to empty each patient_details dictionary :{patient_data[file_id]["patient_details"]})")

            # Adding patient data to Excel
            row = serial_no + 1  # Since row 1 is for headers, data starts from row 2
            # Adding the serial number in the first column
            ws.cell(row=row, column=1, value=serial_no)  

            # Looping through patient_data dictionary and fill each cell in the current row
            for col_num, (key, value) in enumerate(each_person_patient_data.items(), 2):  # starting from column 2
                cell = ws.cell(row=row, column=col_num, value=value)
                # Conditional coloring based on the value in the cell
                if value == "Present":
                    # Green color for "Present"
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif value == "None" or value == "Not Present":
                    # Light Red color for "None" or "Not Present"
                    cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            # Incrementing serial number for the next patient
            serial_no += 1

            # Resetting patient_data dictionary for the next iteration, i'll check afterwards whether it is needed or not.
            # patient_data.clear()

            # Saving the workbook after all data is processed
            wb.save("patient_data.xlsx")
            # It's not needed i guess.
            # patient_data = creating_or_emptying_the_patient_data_dictionary()
            # print(f"This is also the confirmation that the main patient data dictioanary is also emptied : \n {patient_data}")


    elif selected_option == 2:
        # Logic for merged files
        print(f"Input Folder: {input_folder_path}")
        print(f"Output Folder: {output_folder_path}")

        # Looping through the list of PDF files if those are already merged.
        for pdf_file in pdf_files:
            # This will extract the unique key from the file names.
            try:
                original_filename = str(pdf_file).split("\\")[-1]
                file_id = original_filename.split("_")[0].lower()
                if "." in file_id:
                    naming_errors[str(file)] = original_filename
                    print(f"File {pdf_file} has incorrect naming format. Storing naming error: {original_filename}")
                    # Skipping to the next file in the loop, even if there is any naming error also, this will make sure that operations team do thier work properly.
                    continue
                else:
                    if file_id in keys:
                        # If file_id is already in the keys set, add it to the duplicate_file dictionary
                        duplicate_file[file_id] = original_filename
                        print(f"Duplicate file id : {file_id} found in File {pdf_file}, Skipping this file.")
                        # Skipping to the next file in the loop
                        continue
                    else:
                        # Otherwise, adding the file_id to the keys set
                        keys.add(file_id)
            except IndexError:
                original_filename = str(pdf_file).split("\\")[-1]
                naming_errors[str(pdf_file)] = original_filename
                print(f"File {pdf_file} has incorrect naming format. Storing naming error: {original_filename}")
                # Skipping to the next file in the loop, even if there is any naming error also, this will make sure that operations team do thier work properly.
                continue
            print("Keys extracted from file names:", keys)
            print("Naming errors:", naming_errors)

            try:
                # Opening each PDF file in binary mode
                with open(pdf_file, 'rb') as file:
                    # Creating a PdfReader object
                    pdf_reader = PyPDF2.PdfReader(file)
                    # Looping through each page in the PDF and save them as individual PDF files
                    for page_number in range(len(pdf_reader.pages)):
                        # Extract text data from the page to determine the modality
                        page_text = pdf_reader.pages[page_number].extract_text()

                        # Log the page text for debugging, this will print every page.
                        # print_page_text_for_logging(page_text)

                        # This function i've created will check all the conditions and based on that give us the required details.
                        patient_details, modalities = extract_data_based_on_modality(page_text, patient_details, modalities)

                # Now, checking that is there any "None" or empty value in the patient details, if yes , that means there is incomplete data in it.
                missing_keys = [key for key, value in patient_details.items() if value is None]
                if missing_keys:
                    incomplete_data[file_id] = original_filename
                    print(f"Incomplete Data found in file id : {file_id} in File {pdf_file}, Please Review this file.")

                # Checking if patient_id matches with the file_id
                if patient_details['patient_id'] != file_id:
                    id_extracted = patient_details['patient_id']
                    id_mismatch[id_extracted] = original_filename
                    print(f"Id in File : {pdf_file} and in it's filename is not matching, Please Review this file.")
                # Now, I'll update the data in the patien_data dictionary so that i can simply use it to put it in the excel.
                # After processing the pages of the current PDF file, just before moving to the next file:

                # Updating patient_data from patient_details
                for key in patient_details:
                    patient_data[key] = patient_details[key]
                    # if patient_details[key] is not None:
                    #     patient_data[key] = patient_details[key]

                # Checking the modalities set and update corresponding fields in patient_data
                for modality in modalities:
                    if modality in patient_data:
                        patient_data[modality] = 'Present'

                # I'll further process these now, as of now , printing these for additional logs.
                print(f"Patient data for {file_id}: {patient_data}")

                # Clearing the modalities set and patient_details dictionary for the next file
                modalities.clear()
                # Resetting values to None
                for key in patient_details:
                    patient_details[key] = None

                print(f"(This is the confirmation to empty patient_details dictionary :{patient_details})")

                # Adding patient data to Excel
                row = serial_no + 1  # Since row 1 is for headers, data starts from row 2
                # Adding the serial number in the first column
                ws.cell(row=row, column=1, value=serial_no)  

                # Looping through patient_data dictionary and fill each cell in the current row
                for col_num, (key, value) in enumerate(patient_data.items(), 2):  # starting from column 2
                    cell = ws.cell(row=row, column=col_num, value=value)
                    # Conditional coloring based on the value in the cell
                    if value == "Present":
                        # Green color for "Present"
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif value == "None" or value == "Not Present":
                        # Light Red color for "None" or "Not Present"
                        cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

                # Incrementing serial number for the next patient
                serial_no += 1

                # Resetting patient_data dictionary for the next iteration, i'll check afterwards whether it is needed or not.
                # patient_data.clear()

                # Saving the workbook after all data is processed
                wb.save("patient_data.xlsx")

                patient_data = creating_or_emptying_the_patient_data_dictionary()
                print(f"This is also the confirmation that the main patient data dictioanary is also emptied : \n {patient_data}")


            except Exception as e:
                print(f"Error processing {input_folder_path}: {str(e)}")
                exception_files[str(input_folder_path)] = str(e)
                continue  # Skip this file and continue with the next

    # This will be done for both the selections.
    # Save the workbook to the output directory
    output_filename = "Patient_Test_Details.xlsx"
    output_file_path = Path(output_folder_path) / output_filename
    wb.save(output_file_path)

    print(f"All the data extraction is completed and the errors are handled separately, and the data saved successfully to {output_file_path}.")
    
    # Display the completion message
    messagebox.showinfo("Process Completed", 
                        f"Total {num_files_processed} files were Processed.\n\nThe Excel file has been generated and saved to the selected Output Directory : \n{output_file_path}\n\n\nThank you for using OTHM !")
    
    # After processing files, calling the function to handle all errors
    handle_all_errors(naming_errors, duplicate_file, id_mismatch, incomplete_data, exception_files, output_folder_path, input_folder_path)

# These functions, I've created to improve the code readability and reduce the boiler plate codes.
# -------------------------------- HELPER FUNCTIONS (HIMANSHU) --------------------------------------------------------

# This function simply prints the data wnen needed for logging, to make the code non boiler plate code. - Himanshu.
def print_page_text_for_logging(page_text):
    print("Starting of this page(or file) text :")
    print(page_text)
    print("End of this page(or file) text.")

# This function i've made to do both work, creating or emptying the patient data dictionary , when needed.
def creating_or_emptying_the_patient_data_dictionary():
    patient_data = {
            'patient_id': None,
            'patient_name': None,
            'patient_age': None,
            'gender': None,
            'test_date': None,
            'report_date': None,
            'XRAY': 'Not Present',
            'ECG': 'Not Present',
            'AUDIOMETRY': 'Not Present',
            'OPTOMETRY': 'Not Present',
            'VITALS': 'Not Present',
            'PFT': 'Not Present',
            'PATHOLOGY': 'Not Present',
            'OTHERS': 'Not Present'
        }
    return patient_data

# This function i have made to create the patient details dictionary when needed.
def create_patient_details():
    patient_details= {'patient_id': None,'patient_name': None,'patient_age': None,'gender': None,'test_date': None,'report_date': None}
    return patient_details

# This is to ask the user for the input and output path to reduce the code redundancy. - Himanshu.
def select_folders():
    input_folder_path = filedialog.askdirectory(title="Select Input Files Folder", mustexist=True)
    if not input_folder_path:
        print("Input files folder not selected.")
        tk.messagebox.showwarning("Input Directory", "Input directory not selected.")
        return None, None
    
    output_folder_path = filedialog.askdirectory(title="Select Output Files Folder", mustexist=True)
    if not output_folder_path:
        print("Output files folder not selected.")
        tk.messagebox.showwarning("Output Directory", "Output directory not selected.")
        return None, None

    return input_folder_path, output_folder_path

# This function i've created to get the data based on each and every modality.(More optimised way.)
def extract_missing_data_for_modality(modality, page_text, patient_details):
    """
    Extract missing data for the given modality and update the patient details.
    """
    missing_keys = [key for key, value in patient_details.items() if value is None]
    
    if missing_keys:
        # Mapping modalities to their corresponding extraction functions
        modality_functions = {
            'XRAY-BOT': extract_data_from_the_bot_xray_file,
            'XRAY-STRADUS': extract_data_from_the_stradus_xray_file,
            'XRAY-U4RAD-PACS': extract_data_from_the_u4rad_pacs_xray_file,
            'PFT': extract_data_from_bot_pft_file,
            'OPTOMETRY': extract_data_from_bot_opto_file,
            'AUDIOMETRY': extract_data_from_bot_audio_file,
            'ECG': extract_data_from_bot_ecg_file,
            'VITALS': extract_data_from_bot_vitals_file,
            'PATHOLOGY': extract_data_from_the_redcliffe_patho_file
        }

        # This is to Check if the modality exists in the dictionary and extract the data
        # This will check the respective key(which is the modality here, which i pass while calling the function.)
        # The key takeaway is the modality i pass and the modality i add in the set can be different, so i can map these different functions like 
        # Extracting the data from the stradus / our pacs report etc. the function will set the modality_data as an empty dictionary, when no modality is found, so i can handle any part here now.
        # More details if needed than see in my documentation or search on web.
        modality_data = modality_functions.get(modality, lambda x: {}) (page_text)
        
        # Logging the extraction message
        if modality_data:
            # This will tell me the patient details when any field in it is empty.
            print(f"This is the patient details just before adding the extracted data in it, because some fields were empty in it :\n {patient_details}")
            print(f"Data extracted from the {modality} report.")
        
        # Updating patient details with the extracted data
        for key in missing_keys:
            if key in modality_data:
                patient_details[key] = modality_data[key]

    return patient_details

from openpyxl import Workbook
from openpyxl.styles import Font

from openpyxl import Workbook
from openpyxl.styles import Font

def modality_based_excel_workbook(modality):
    print(f"This is the modality :{modality}")
    # Headers for each modality
    headers = {
        'XRAY': ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE', 'FINDINGS'],
        'PFT': ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE', 'HEIGHT', 'WEIGHT', 'OBSERVATIONS'],
        'OPTOMETRY': ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE',
                      'FAR VISION RIGHT', 'NEAR VISION RIGHT', 'DISTANCE VISION RIGHT', 'READING RIGHT','SPHERICAL RIGHT', 
                      'CYLINDRICAL RIGHT', 'AXIS RIGHT', 'ADD RIGHT',
                      'FAR VISION LEFT', 'NEAR VISION LEFT', 'DISTANCE VISION LEFT', 'READING VISION LEFT', 
                      'SPHERICAL LEFT', 'CYLINDRICAL LEFT', 'AXIS LEFT', 'ADD LEFT', 'COLOUR BLINDNESS'],
        'AUDIOMETRY': ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE', 'LEFT EAR FINDING', 'RIGHT EAR FINDING'],
        'ECG': ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE', 'HEART RATE', 'FINDINGS'],
        'VITALS': ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE', 'HEIGHT', 'WEIGHT', 'BLOOD PRESSURE', 'PULSE'],
        'PATHOLOGY': ['SERIAL NO.', 'PATIENT ID', 'PATIENT NAME', 'AGE', 'GENDER', 'STUDY DATE', 'REPORT DATE', 
                      'HAEMOGLOBIN', 'RBC COUNT', 'PCV', 'MCV', 'MCH', 'MCHC', 'RDW (CV)', 'RDW-SD',
                      'TLC', 'NEUTROPHILS (DLC)', 'LYMPHOCYTES (DLC)', 'MONOCYTES (DLC)', 'EOSINOPHILS (DLC)', 'BASOPHILS (DLC)', 
                      'NEUTROPHILS (ALC)', 'LYMPHOCYTES (ALC)', 'MONOCYTES (ALC)', 'EOSINOPHILS (ALC)', 'BASOPHILS (ALC)', 
                      'PLATELET COUNT', 'MPV', 'PCT']
    }

    # Always create a new workbook for each modality
    modality_workbook = {
        'XRAY': Workbook(),
        'PFT': Workbook(),
        'OPTOMETRY': Workbook(),
        'AUDIOMETRY': Workbook(),
        'ECG': Workbook(),
        'VITALS': Workbook(),
        'PATHOLOGY': Workbook()
    }

    modality_patient_data_dictionary = {
        key: {header: None for header in headers[key]} for key in headers
    }

    # Get the workbook for the requested modality
    wb = modality_workbook.get(modality)
    if wb is None:
        print(f"Error: Workbook for modality '{modality}' not found!")
        return None, None, None

    ws = wb.active  # get first worksheet
    modality_headers = headers.get(modality)
    respective_patient_data = modality_patient_data_dictionary.get(modality)

    if wb and modality_headers:
        # ✅ Corrected line
        if not any(cell.value for cell in ws[1]):  
            # Add headers
            for col_num, header in enumerate(modality_headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
            serial_no = 1
        else:
            serial_no = ws.max_row + 1

        return wb, serial_no, respective_patient_data
    else:
        return None, None, None



# This function i've created to check or get the data based on conditions.
def extract_data_based_on_modality(page_text, patient_details, modalities):
    """
    Function to check for different modalities in the page_text, 
    extract missing data, and update patient details accordingly.
    
    Parameters:
    - page_text (str): Text extracted from the page
    - patient_details (dict): A dictionary containing patient details
    - modalities (set): A set to store the modalities found
    
    Returns:
    - updated_patient_details (dict): The updated patient details with newly extracted data
    - updated_modalities (set): The updated set of modalities found
    """
    # Define unwanted phrases
    unwanted_phrases = ["Referrer Dr.", "left ear", "RECORDERS & MEDICARE SYSTEMS", 
                        "OPTOMETRY", "ECG", "VITALS", "RBC Count", "PDW *", "PDW"]

    # Check the text content for known modalities
    # For the xray report from stradus.
    if "Referrer Dr" in page_text and "Time" in page_text:
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        modalities.add('XRAY')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('XRAY-STRADUS', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    # For the xray report from u4rad pacs reporting bot.
    if "Test Date:" in page_text and "Report Date:" in page_text and not any(phrase in page_text for phrase in unwanted_phrases):
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        modalities.add('XRAY')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('XRAY-U4RAD-PACS', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    # For the xray report from reporting bot.
    if "X-RAY" in page_text:
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        modalities.add('XRAY')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('XRAY-BOT', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    elif "RECORDERS & MEDICARE SYSTEMS" in page_text:
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        print("This is a PFT file.")
        modalities.add('PFT')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('PFT', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    elif "OPTOMETRY" in page_text:
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        print("This is an Optometry file.")
        modalities.add('OPTOMETRY')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('OPTOMETRY', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    elif "left ear" in page_text:
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        print("This is an Audiometry file.")
        modalities.add('AUDIOMETRY')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('AUDIOMETRY', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    elif "ECG" in page_text:
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        print("This is an ECG file.")
        modalities.add('ECG')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('ECG', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    elif page_text == '':
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        print("This is an Others image.")
        print("Since it is an others file, means the page is blank, No data to extract.")
        modalities.add('OTHERS')
    elif "VITALS" in page_text:
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        print("This is a Vitals file.")
        modalities.add('VITALS')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('VITALS', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    elif "RBC Count" in page_text or "PDW *" in page_text or "PDW" in page_text :
        # This will print text only when any modality matches.
        print_page_text_for_logging(page_text)
        print("This is a Blood Report.")
        modalities.add('PATHOLOGY')
        # Calling the data extractor function i've created wrt each conditional expectation.
        patient_details = extract_missing_data_for_modality('PATHOLOGY', page_text, patient_details)
        # Additional logging to see the patient details extracted so far, and it will give the patient details only if some thing is missing or empty.
        print(f"This is the patient details after adding the data which we extracted and because the details were not completely filled before :\n {patient_details}")

    # Only the above conditions will tell that whether it is a particular test or not, if others are needed , I'll update them here.


    return patient_details, modalities


# This is my function to extract data from all the pft files :
def extract_data_from_bot_pft_file(pageText):
    # Extracting the required information from the text
    patient_info = {}
    
    try:
        patient_info['patient_id'] = str(pageText).split("ID")[1].split("Age")[0].split(":")[1].strip()
        patient_info['patient_name'] = str(pageText).split("Patient")[1].split("Refd.By:")[0].split(":")[1].strip()
        patient_info['patient_age'] = str(pageText).split("Age    :")[1].split("Yrs")[0].strip()
        patient_info['gender'] = str(pageText).split("Gender")[1].split("Smoker")[0].split(":")[1].strip()
        patient_info['height'] = str(pageText).split("Height :")[1].split("Weight")[0].strip()
        patient_info['weight'] = str(pageText).split("Weight")[1].split("Gender")[0].split(":")[1].split("Kgs")[0].strip()
        patient_info['test_date'] = str(pageText).split("Date")[1][1:21].split(":")[1]
        patient_info['report_date'] = patient_info['test_date']
        patient_info['observation'] = str(pageText).split("Pre Test COPD Severity")[1].strip()
    except IndexError:
        print("Error extracting data from the first page text.")
    
    return patient_info

# This is my function for extracting data from all ECG reports:
def extract_data_from_bot_ecg_file(pageText):
    patient_info = {}
    try:
        patient_info['patient_id'] = str(pageText).split('Patient ID:')[1].split('Age:')[0].strip()
        patient_info['patient_name'] = str(pageText).split("Name:")[1].split("Patient ID:")[0].strip()
        patient_info['patient_age'] = str(pageText).split("Age:")[1].split('Gender:')[0].strip()
        patient_info['gender'] = str(pageText).split("Gender:")[1].split("Test date:")[0].strip()
        patient_info['test_date'] = str(pageText).split("Test date:")[1].split('Report date:')[0].strip()
        patient_info['report_date'] = str(pageText).split("Report date:")[1].split('ECG')[0].strip()
        patient_info['heart_rate'] = str(pageText).split("Heart rate is")[1].split("BPM.")[0].strip()
        patient_info['findings'] = str(pageText).split("2.")[1].split('.')[0].strip()
    except IndexError:
        print("Error extracting ECG data.")
    return patient_info

# This is my function for extracting data from X-ray reports of our Reporting Bot:
def extract_data_from_the_bot_xray_file(pageText):
    patient_info = {}
    try:
        # Check if the "IMPRESSION:" text exists for finding data, right now , i am commenting it.
        if 'IMPRESSION:' in pageText:
            findings_data = str(pageText).split('IMPRESSION:')[1].split("Dr.")[0]
            if "•" in findings_data:
                findings = findings_data.split("•")[1].split(".")[0].strip()
            else:
                findings = findings_data.strip()
        else:
            findings = None
        
        # If specific "Study Date" and "Report Date" condition applies, this i've still included in case the format changes.
        if "Study Date" and "Report Date" in pageText:
            patient_info['patient_id'] = str(pageText).split("Patient ID")[1].split(" ")[1].lower().strip()
            patient_info['patient_name'] = str(pageText).split("Name")[1].split("Date")[0].split(" ")[0].strip().lower()
            if "patient" in patient_info['patient_name']:
                patient_info['patient_name'] = patient_info['patient_name'].split("patient")[0].strip()
            print(patient_info['patient_id'], patient_info['patient_name'])
        # Mostly this will be the case for our bot generated xrays.
        else:
            patient_info['patient_id'] = str(pageText).split('Patient ID:')[1].split('Age:')[0].strip()
            patient_info['patient_name'] = str(pageText).split('Name:')[1].split('Patient ID:')[0].strip()
            patient_info['patient_age'] = str(pageText).split('Age:')[1].split('Gender:')[0].strip()
            patient_info['gender'] = str(pageText).split('Gender:')[1].split('Test date:')[0].strip()
            patient_info['test_date'] = str(pageText).split('Test date:')[1].split('Report date:')[0].strip()
            patient_info['report_date'] = str(pageText).split('Report date:')[1].split('X-RAY')[0].strip()
        # This i will use to clean up the findings of the data.
        patient_info['findings'] = remove_illegal_characters(findings)
    except IndexError:
        print("Error extracting X-ray data.")
    return patient_info

# This is my function for extracting data from Stradus X-ray reports:
def extract_data_from_the_stradus_xray_file(pageText):
    patient_info = {}
    try:
        patient_info['patient_id'] = str(pageText).split('Patient ID')[1].split('Age')[0].strip()
        patient_info['patient_name'] = str(pageText).split('Patient Name')[1].split('Patient ID')[0].strip()
        patient_info['patient_age'] = 'Non-Extractable'
        patient_info['gender'] = str(pageText).split('Sex')[1].split('Study Date')[0].strip()
        patient_info['test_date'] = str(pageText).split('Study Date')[1].split('Report Date')[0].split('Time')[1].split('\n')[0].strip()
        # It is really varying here, sometimes there is "impression:", sometimes "IMPRESSION", similarly for word "observations" , and sometimes these words are even not present there.
        report_date_data = str(pageText).split('Report Date')[1].split('Dr.')[0]
        # Considering the case when there is am in time.
        if 'am' in report_date_data:
            patient_info['report_date'] = str(pageText).split('Report Date')[1].split('am')[0].split('Time')[1].split('\n')[0].strip()
        # Considering the case, when there is pm in time.
        else:
            patient_info['report_date'] = str(pageText).split('Report Date')[1].split('pm')[0].split('Time')[1].split('\n')[0].strip()
        
        # I will extract other data from here afterwards.
        # I know that the doctors will make this mistake so i'm fixing it here (maximum cases). - Himanshu.
        if 'IMPRESSION :-' in pageText:
            findings_data = str(pageText).split('IMPRESSION :-')[1].split("ADVICE :-")[0]
            if "•" in findings_data:
                findings = findings_data.split("•")[1].split(".")[0].strip()
            else:
                findings = findings_data.strip()
        elif 'IMPRESSIONS :-' in pageText:
            findings_data = str(pageText).split('IMPRESSIONS :-')[1].split("ADVICE :-")[0]
            if "•" in findings_data:
                findings = findings_data.split("•")[1].split(".")[0].strip()
            else:
                findings = findings_data.strip()
        else:
            findings = None
        # Now, adding the findings in the excel.
        patient_info['findings'] = remove_illegal_characters(findings)
        
    except IndexError:
        print("Error extracting X-ray data.")
    return patient_info

# This is my function for extracting data from our orthanc pacs X-ray reports:
def extract_data_from_the_u4rad_pacs_xray_file(pageText):
    patient_info = {}
    try:
        patient_info['patient_id'] = str(pageText).split('Patient ID:')[1].split('Patient Age:')[0].strip()
        patient_info['patient_name'] = str(pageText).split('Patient Name:')[1].split('Patient ID:')[0].strip()
        patient_info['patient_age'] = str(pageText).split('Patient Age:')[1].split('Patient Gender:')[0].strip()
        patient_info['gender'] = str(pageText).split('Patient Gender:')[1].split('Test Date:')[0].strip()
        patient_info['test_date'] = str(pageText).split('Test Date:')[1].split('Report Date:')[0].strip()
        patient_info['report_date'] = str(pageText).split('Report Date:')[1].split('Dr.')[0].split('\n')[0].strip()
        
        # Extract findings based on known formats
        findings = None
        for key in ['IMPRESSION:', 'IMPRESSIONS:', 'IMPRESSION;', 'IMPRESSIONS;', 'IMPRESSION :-', 'ADVICE :-']:
            if key in pageText:
                findings_data = pageText.split(key)[1].split("Dr.")[0]
                if "•" in findings_data:
                    findings = findings_data.split("•")[1].split(".")[0].strip()
                else:
                    findings = findings_data.strip()
                break

        # Default to standard phrase if findings is None or empty
        if not findings or not findings.strip():
            findings = " No significant abnormality detected."  # Ensure period

        # Sanitize and ensure period consistency
        findings = findings.strip()
        if not findings.endswith('.'):
            findings += '.'
        patient_info['findings'] = remove_illegal_characters(findings)

        # Sanitize findings and add to patient info
        patient_info['findings'] = remove_illegal_characters(findings.strip())

    except IndexError:
        print("Error extracting X-ray data.")
    
    return patient_info


# Function for extracting data from Blood reports
def extract_data_from_the_redcliffe_patho_file(pageText):
    patient_info = {}
    try:
        # Handling the extraction of patient details with error handling for each field
        try:
            if "Patient Name :" in pageText:
                complete_patient_name = str(pageText).split("Patient Name : ")[1].split("DOB/")[0].strip()
                patient_info['patient_id'] = complete_patient_name.rsplit(" ", 1)[1]
                patient_info['patient_name'] = complete_patient_name.rsplit(" ", 1)[0].split(" ", 1)[1].lower()
            elif "Patient NAME :" in pageText:
                complete_patient_name = str(pageText).split("Patient NAME : ")[1].split("DOB/")[0].strip()
                patient_info['patient_id'] = complete_patient_name.rsplit("_", 1)[1]
                patient_info['patient_name'] = complete_patient_name.rsplit("_", 1)[0].split(" ", 1)[1].lower()
            elif "PATIENT NAME :" in pageText:
                complete_patient_name = str(pageText).split("PATIENT NAME : ")[1].split("DOB/")[0].strip()
                patient_info['patient_id'] = complete_patient_name.rsplit("_", 1)[1]
                patient_info['patient_name'] = complete_patient_name.rsplit("_", 1)[0].split(" ", 1)[1].lower()
            elif "PATIENT Name :" in pageText:
                complete_patient_name = str(pageText).split("PATIENT Name : ")[1].split("DOB/")[0].strip()
                patient_info['patient_id'] = complete_patient_name.rsplit("_", 1)[1]
                patient_info['patient_name'] = complete_patient_name.rsplit("_", 1)[0].split(" ", 1)[1].lower()
            else:
                patient_info['patient_id'] = None
                patient_info['patient_name'] = None
        except Exception as e:
            patient_info['patient_id'] = None
            patient_info['patient_name'] = None
            print(f"Error extracting patient details: {e}")
        
        # Age, gender, and test dates with error handling
        try:
            patient_info['patient_age'] = str(pageText).split('DOB/Age/Gender :')[1].split('Patient ID / UHID :')[0].split('Y/')[0].strip()
            patient_info['gender'] = str(pageText).split('DOB/Age/Gender :')[1].split('Patient ID / UHID :')[0].split('Y/')[1].strip()
            patient_info['test_date'] = str(pageText).split('Sample Collected :')[1].split('Report STATUS :')[0].strip()
            patient_info['report_date'] = str(pageText).split('Report Date :')[1].split('Test Description')[0].strip()
        except Exception as e:
            patient_info['patient_age'] = None
            patient_info['gender'] = None
            patient_info['test_date'] = None
            patient_info['report_date'] = None
            print(f"Error extracting age, gender, or dates: {e}")
        
        # Hemoglobin extraction with error handling
        try:
            if 'Hemoglobin' in pageText:
                patient_info['haemoglobin'] = str(pageText).split('Hemoglobin')[1].split('colorimetric')[1].split(' ')[0].strip()
            elif 'Haemoglobin' in pageText:
                patient_info['haemoglobin'] = str(pageText).split('Haemoglobin')[1].split('colorimetric')[1].split(' ')[0].strip()
            else:
                patient_info['haemoglobin'] = None
        except Exception as e:
            patient_info['haemoglobin'] = None
            print(f"Error extracting hemoglobin: {e}")
        
        # RBC, WBC, Platelet counts with error handling
        try:
            patient_info['rbc_count'] = str(pageText).split('RBC Count')[1].split('Electrical impedance')[1].split(' ')[0].strip()
            patient_info['rbc_pcv'] = str(pageText).split('PCV')[1].split('Calculated')[1].split(' ')[0].strip()
            patient_info['rbc_mcv'] = str(pageText).split('MCV')[1].split('Calculated')[1].split(' ')[0].strip()
            patient_info['rbc_mch'] = str(pageText).split('MCH')[1].split('Calculated')[1].split(' ')[0].strip()
            patient_info['rbc_mchc'] = str(pageText).split('MCHC')[1].split('Calculated')[1].split(' ')[0].strip()
            patient_info['rbc_rdw_cv'] = str(pageText).split('RDW (CV)')[1].split('Calculated')[1].split(' ')[0].strip()
            patient_info['rbc_rdw_sd'] = str(pageText).split('RDW-SD')[1].split('Calculated')[1].split(' ')[0].strip()
            patient_info['wbc_tlc'] = str(pageText).split('TLC')[1].split('Electrical impedance and microscopy')[1].split(' ')[0].strip()
        except Exception as e:
            patient_info['rbc_count'] = None
            patient_info['rbc_pcv'] = None
            patient_info['rbc_mcv'] = None
            patient_info['rbc_mch'] = None
            patient_info['rbc_mchc'] = None
            patient_info['rbc_rdw_cv'] = None
            patient_info['rbc_rdw_sd'] = None
            patient_info['wbc_tlc'] = None
            print(f"Error extracting RBC/WBC counts: {e}")
        
        # Differential counts with error handling
        try:
            patient_info['dlc_neutrophils'] = str(pageText).split('Differential Leucocyte Count')[1].split('Neutrophils')[1].split(' ')[0].strip()
            patient_info['dlc_lymphocytes'] = str(pageText).split('Differential Leucocyte Count')[1].split('Lymphocytes')[1].split(' ')[0].strip()
            patient_info['dlc_monocytes'] = str(pageText).split('Differential Leucocyte Count')[1].split('Monocytes')[1].split(' ')[0].strip()
            patient_info['dlc_eosinophils'] = str(pageText).split('Differential Leucocyte Count')[1].split('Eosinophils')[1].split(' ')[0].strip()
            patient_info['dlc_basophils'] = str(pageText).split('Differential Leucocyte Count')[1].split('Basophils')[1].split(' ')[0].strip()
        except Exception as e:
            patient_info['dlc_neutrophils'] = None
            patient_info['dlc_lymphocytes'] = None
            patient_info['dlc_monocytes'] = None
            patient_info['dlc_eosinophils'] = None
            patient_info['dlc_basophils'] = None
            print(f"Error extracting DLC counts: {e}")
        
        # Absolute Leukocyte Counts with error handling
        try:
            patient_info['alc_neutrophils'] = str(pageText).split('Absolute Leukocyte Counts')[1].split('Neutrophils.')[1].split(' ')[0].strip()
            patient_info['alc_lymphocytes'] = str(pageText).split('Absolute Leukocyte Counts')[1].split('Lymphocytes.')[1].split(' ')[0].strip()
            patient_info['alc_monocytes'] = str(pageText).split('Absolute Leukocyte Counts')[1].split('Monocytes.')[1].split(' ')[0].strip()
            patient_info['alc_eosinophils'] = str(pageText).split('Absolute Leukocyte Counts')[1].split('Eosinophils.')[1].split(' ')[0].strip()
            patient_info['alc_basophils'] = str(pageText).split('Absolute Leukocyte Counts')[1].split('Basophils.')[1].split(' ')[0].strip()
        except Exception as e:
            patient_info['alc_neutrophils'] = None
            patient_info['alc_lymphocytes'] = None
            patient_info['alc_monocytes'] = None
            patient_info['alc_eosinophils'] = None
            patient_info['alc_basophils'] = None
            print(f"Error extracting ALC counts: {e}")
        
        # Platelet counts and other tests with error handling
        try:
            patient_info['platelet_count'] = str(pageText).split('Platelet Count')[1].split('Electrical impedance and microscopy')[1].split(' ')[0].strip()
            patient_info['mean_platelet_volume'] = str(pageText).split('Mean Platelet Volume (MPV)')[1].split('Calculated')[1].split(' ')[0].strip()
            if 'PCT' in pageText:
                patient_info['pct'] = str(pageText).split('PCT')[1].split('Calculated')[1].split(' ')[0].strip()
            else:
                patient_info['pct'] = None
        except Exception as e:
            patient_info['platelet_count'] = None
            patient_info['mean_platelet_volume'] = None
            patient_info['pct'] = None
            print(f"Error extracting platelet data: {e}")

        print("Patient Name:", patient_info['patient_name'])
        print("Patient ID:", patient_info['patient_id'])

    except Exception as e:
        print(f"General error in extracting data: {e}")

    return patient_info


# This is my function for extracting data from all Reporting Bot Audiometry Reports:
def extract_data_from_bot_audio_file(pageText):
    patient_info = {}
    try:
        patient_info['patient_id'] = str(pageText).split('Patient ID')[1].split('Age')[0].strip()
        patient_info['patient_name'] = str(pageText).split("Name")[1].split("Patient ID")[0].strip()
        patient_info['patient_age'] = str(pageText).split("Age")[1].split('Gender')[0].strip()
        patient_info['gender'] = str(pageText).split("Gender")[1].split("Test date")[0].strip()
        patient_info['test_date'] = str(pageText).split("Test date")[1].split('Report date')[0].strip()
        patient_info['report_date'] = str(pageText).split("Report date")[1].strip()
        patient_info['left_ear_finding'] = str(pageText).split("in left ear")[0].split("Finding:")[1].strip()
        patient_info['right_ear_finding'] = str(pageText).split("in right ear")[0].split("in left ear.")[1].strip()
    except IndexError:
        print("Error extracting Audiometry data.")
    return patient_info

# This is my function for extracting data from all Reporting Bot Optometry Reports:
def extract_data_from_bot_opto_file(pageText):
    patient_info = {}
    try:
        patient_info['patient_id'] = str(pageText).split('Patient ID:')[1].split('Patient Name:')[0].strip()
        patient_info['patient_name'] = str(pageText).split("Patient Name:")[1].split("Age:")[0].strip()
        patient_info['patient_age'] = str(pageText).split("Age:")[1].split('Gender:')[0].strip()
        patient_info['gender'] = str(pageText).split("Gender:")[1].split("Test Date:")[0].strip()
        patient_info['test_date'] = str(pageText).split("Test Date:")[1].split('Report Date:')[0].strip()
        patient_info['report_date'] = str(pageText).split("Report Date:")[1].split('OPTOMETRY')[0].strip()
        patient_info['far_vision_right'] = str(pageText).split("Distance(Far):")[1].split('vision in right eye')[0].strip()
        patient_info['near_vision_right'] = str(pageText).split("Distance(Near):")[1].split('vision in right eye')[0].strip()
        patient_info['distance_vision_right'] = str(pageText).split("Right Eye")[1].split('Left Eye')[0].strip().split(" ")[0].strip()
        patient_info['reading_vision_right'] = str(pageText).split("Right Eye")[1].split('Left Eye')[0].strip().split(" ")[1].strip()
        patient_info['spherical_right'] = str(pageText).split("Right Eye")[1].split('Left Eye')[0].strip().split(" ")[2].strip()
        patient_info['cylindrical_right'] = str(pageText).split("Right Eye")[1].split('Left Eye')[0].strip().split(" ")[3].strip()
        patient_info['axis_right'] = str(pageText).split("Right Eye")[1].split('Left Eye')[0].strip().split(" ")[4].strip()
        patient_info['add_right'] = str(pageText).split("Right Eye")[1].split('Left Eye')[0].strip().split(" ")[5].strip()
        patient_info['far_vision_left'] = str(pageText).split("Distance(Far):")[1].split('vision in right eye-')[1].split("vision in left eye.")[0].strip()
        patient_info['near_vision_left'] = str(pageText).split("Distance(Far):")[1].split('vision in right eye-')[1].split("vision in left eye.")[0].strip()
        patient_info['distance_vision_right'] = str(pageText).split("Left Eye")[1].split('Color vision')[0].strip().split(" ")[0].strip()
        patient_info['reading_vision_left'] = str(pageText).split("Left Eye")[1].split('Color vision')[0].strip().split(" ")[1].strip()
        patient_info['spherical_left'] = str(pageText).split("Left Eye")[1].split('Color vision')[0].strip().split(" ")[2].strip()
        patient_info['cylindrical_left'] = str(pageText).split("Left Eye")[1].split('Color vision')[0].strip().split(" ")[3].strip()
        patient_info['axis_left'] = str(pageText).split("Left Eye")[1].split('Color vision')[0].strip().split(" ")[4].strip()
        patient_info['add_left'] = str(pageText).split("Left Eye")[1].split('Color vision')[0].strip().split(" ")[5].strip()
        patient_info['colour_blindness'] = str(pageText).split("Color vision check(Ishihara test):")[1].split('color blindness')[0].strip()
    except IndexError:
        print("Error extracting Optometry data.")
    return patient_info

# This is my function for extracting data from all Reporting Bot Vitals Reports:
def extract_data_from_bot_vitals_file(pageText):
    patient_info = {}
    try:
        # This is for our reporting bot vitals file.
        if "Patient Name:" in pageText:
            patient_info['patient_id'] = str(pageText).split('Patient ID:')[1].split('Patient Name:')[0].strip()
            patient_info['patient_name'] = str(pageText).split("Patient Name:")[1].split("Age:")[0].strip()
            patient_info['patient_age'] = str(pageText).split("Age:")[1].split('Gender:')[0].strip()
            patient_info['gender'] = str(pageText).split("Gender:")[1].split("Test Date:")[0].strip()
            patient_info['test_date'] = str(pageText).split("Test Date:")[1].split('Report Date:')[0].strip()
            patient_info['report_date'] = str(pageText).split("Report Date:")[1].split('VITALS')[0].strip()
        # This was some other format which i found in one of the vitals files.
        else:
            patient_info['patient_id'] = str(pageText).split('Patient ID:')[1].split('Age:')[0].strip()
            patient_info['patient_name'] = str(pageText).split("Name:")[1].split("Patient ID:")[0].strip()
            patient_info['patient_age'] = str(pageText).split("Age:")[1].split('Gender:')[0].strip()
            patient_info['gender'] = str(pageText).split("Gender:")[1].split("Test date:")[0].strip()
            patient_info['test_date'] = str(pageText).split("Test date:")[1].split('Report date:')[0].strip()
            patient_info['report_date'] = str(pageText).split("Report date:")[1].split('VITALS')[0].strip()
            patient_info['height'] = str(pageText).split("Height(in cm)")[1].split("Weight(in kg)")[0].strip()
            patient_info['weight'] = str(pageText).split("Weight(in kg)")[1].split("BMI(kg/m2)")[0].strip()
            patient_info['bp'] = str(pageText).split("Blood Pressure(mmHg)")[1].split("Pulse(bpm)")[0].strip()
            patient_info['Pulse'] = str(pageText).split("Pulse(bpm)")[1].strip()
    except IndexError:
        print("Error extracting Vitals data.")
    return patient_info

# Function to write errors to the error file and accumulate the messages
import shutil
from pathlib import Path

def write_errors_to_file(naming_errors, duplicate_file, id_mismatch, incomplete_data, exception_files, output_folder_path, input_folder_path):
    # Ensure output_folder_path exists
    output_folder_path = Path(output_folder_path)
    output_folder_path.mkdir(parents=True, exist_ok=True)

    # Converting input folder path also into path object.
    input_folder_path = Path(input_folder_path)
    
    # Define the paths for the error subdirectories
    naming_error_folder = output_folder_path / "NamingErrorFiles"
    duplicate_files_folder = output_folder_path / "DuplicateFiles"
    id_mismatch_folder = output_folder_path / "IdMismatchFiles"
    incomplete_data_folder = output_folder_path / "IncompleteDataFiles"
    exception_files_folder = output_folder_path / "ExceptionFiles"

    # Define the error details file path
    error_details_file = output_folder_path / "ErrorDetails.txt"
    
    with open(error_details_file, "w") as file:
        file.write("The files that were processed were having the following errors:\n")
        file.write("================================================================\n\n")

        # Naming Errors
        if naming_errors:
            naming_error_count = len(naming_errors)
            file.write(f"NAMING ERRORS:\n\n")
            file.write(f"{naming_error_count} file(s) is/are having naming issues, so skipping them from processing :\nNOTE: You should follow A_B* naming convention in filename where 'A' is id and 'B' is considered as name \n\n")
            for file_id, original_filename in naming_errors.items():
                file.write(f"Filename: {original_filename}\n")
                # Copy the file to the respective folder
                naming_error_folder.mkdir(parents=True, exist_ok=True)
                # making folder only when there is respective error.
                try:
                    original_file_path = input_folder_path / original_filename
                    shutil.copy2(original_file_path, naming_error_folder / original_filename)
                except Exception as e:
                    file.write(f"Error copying {original_filename} to NamingErrorFiles: {e}\n")
                    file.write("Please Contact Himanshu to resolve this issue.\n")
            file.write("\n-------------------------------\n\n")

        # Duplicate Files
        if duplicate_file:
            duplicate_file_count = len(duplicate_file)
            file.write(f"DUPLICATE FILES:\n\n")
            file.write(f"{duplicate_file_count} file(s) is/are duplicates, so skipping it from processing :\n\n")
            for file_id, original_filename in duplicate_file.items():
                file.write(f"File ID: {file_id}, Filename: {original_filename}\n")
                # Copy the file to the respective folder
                duplicate_files_folder.mkdir(parents=True, exist_ok=True)
                # making folder only when there is respective error.
                try:
                    original_file_path = input_folder_path / original_filename
                    shutil.copy2(original_file_path, duplicate_files_folder / original_filename)
                except Exception as e:
                    file.write(f"Error copying {original_filename} to DuplicateFiles: {e}\n")
                    file.write("Please Contact Himanshu to resolve this issue.\n")
            file.write("\n-------------------------------\n\n")

        # ID Mismatch
        if id_mismatch:
            id_mismatch_count = len(id_mismatch)
            file.write(f"ID MISMATCH:\n\n")
            file.write(f"{id_mismatch_count} file(s) have ID mismatches, still it's included in processing :\nThe id in file and in filename is not matching.\n\n")
            for file_id, original_filename in id_mismatch.items():
                file.write(f"ID in File: {file_id}, Filename: {original_filename}\n")
                # Copy the file to the respective folder
                id_mismatch_folder.mkdir(parents=True, exist_ok=True)
                # making folder only when there is respective error.
                try:
                    original_file_path = input_folder_path / original_filename
                    shutil.copy2(original_file_path, id_mismatch_folder / original_filename)
                except Exception as e:
                    file.write(f"Error copying {original_filename} to IdMismatchFiles: {e}\n")
                    file.write("Please Contact Himanshu to resolve this issue.\n")
            file.write("\n-------------------------------\n\n")

        # Incomplete Data
        if incomplete_data:
            incomplete_data_count = len(incomplete_data)
            file.write(f"INCOMPLETE DATA FILES:\n\n")
            file.write(f"{incomplete_data_count} file(s) have incomplete data :\n\n")
            for file_id, filename in incomplete_data.items():
                file.write(f"File ID: {file_id}, Filename: {filename}\n")
                # Copy the file to the respective folder
                incomplete_data_folder.mkdir(parents=True, exist_ok=True)
                # making folder only when there is respective error.
                try:
                    original_file_path = input_folder_path / filename
                    shutil.copy2(original_file_path, incomplete_data_folder / filename)
                except Exception as e:
                    file.write(f"Error copying {filename} to IncompleteDataFiles: {e}\n")
                    file.write("Please Contact Himanshu to resolve this issue.\n")
            file.write("\n-------------------------------\n\n")

        # Exception Files
        if exception_files:
            exception_file_count = len(exception_files)
            file.write(f"EXCEPTION FILES:\n\n")
            file.write(f"{exception_file_count} file(s) encountered errors :\n\n")
            for filename, error_message in exception_files.items():
                file.write(f"FileName: {filename}, Error: {error_message}\n")
                # Copy the file to the respective folder
                exception_files_folder.mkdir(parents=True, exist_ok=True)
                # making folder only when there is respective error.
                try:
                    original_file_path = input_folder_path / filename
                    shutil.copy2(original_file_path, exception_files_folder / filename)
                except Exception as e:
                    file.write(f"Error copying {filename} to ExceptionFiles: {e}\n")
                    file.write("Please Contact Himanshu to resolve this issue.\n")
            file.write("\n-------------------------------\n\n")

    # I can make the above code more optimized where i am using multiple times the file.write functionality.
    # I will update it later.
    
    print(f"Error details written to {error_details_file}")


# Show warning message box if there are any errors
def show_warning_message(naming_errors, duplicate_file, id_mismatch, incomplete_data, exception_files):
    warning_message = ""

    if naming_errors:
        warning_message += f"{len(naming_errors)} file(s) had naming conflicts.\n"
    if duplicate_file:
        warning_message += f"{len(duplicate_file)} duplicate file(s) found.\n"
    if id_mismatch:
        warning_message += f"{len(id_mismatch)} file(s) had ID mismatches.\n"
    if incomplete_data:
        warning_message += f"{len(incomplete_data)} file(s) had incomplete data.\n"
    if exception_files:
        warning_message += f"{len(exception_files)} problematic file(s) encountered.\n"

    if warning_message:
        tk.messagebox.showwarning("Errors in File Processing", warning_message)

# Main function to handle all errors and generate the file
def handle_all_errors(naming_errors, duplicate_file, id_mismatch, incomplete_data, exception_files, output_folder_path, input_folder_path):
    if naming_errors or duplicate_file or id_mismatch or incomplete_data or exception_files:
        # Write all errors to the error details file
        write_errors_to_file(naming_errors, duplicate_file, id_mismatch, incomplete_data, exception_files, output_folder_path, input_folder_path)
        # Show the warning message box with a consolidated summary of errors
        show_warning_message(naming_errors, duplicate_file, id_mismatch, incomplete_data, exception_files)
    else:
        print("No errors found.")


# This function i will make later , when i will include our orthanc pacs generated xray (or all) reports in automations.

#  ----------------------------------- END OF HELPER FUNCTIONS (HIMANSHU) ------------------------------------------------------------------

# def dcm_to_pdf_converter():
#     input_directory = filedialog.askdirectory(title="Select Input Directory")
#
#     if input_directory:
#         # Prompt user to select output directory
#         output_directory = filedialog.askdirectory(title="Select Output Directory")
#
#         if output_directory:
#             input_dir = Path(input_directory)
#             output_dir = Path(output_directory)
#             output_dir.mkdir(parents=True, exist_ok=True)
#
#             error_dir = output_dir / "error_files"
#             error_dir.mkdir(parents=True, exist_ok=True)
#
#             pdf_files = list(input_dir.glob("*.pdf"))
#             error_count = 0
#
#
#
#         else:
#             messagebox.showwarning("Output Folder Not Selected", "Output folder not selected.")
#     else:
#         messagebox.showwarning("Input Folder Not Selected", "Input folder not selected.")

# Create the main window
window = tk.Tk()
window.title("OTHM - Operation Tasks Helping Machine")
# Set the window dimensions and position it on the screen
window.geometry("1000x500+200-100")


redcliffe_label = tk.Label(window, text="Merging For Individual", font=("Arial", 16, "bold"))
redcliffe_label.place(x=620, y=10, anchor='ne')

# Adding the label of Merge All files button .
merge_all_files = tk.Label(window, text="Merge Everything", font=("Arial", 16, "bold"))
merge_all_files.place(x=600, y=130, anchor='ne')

merge_redcliffe_button1 = tk.Button(window, bg='blue', fg='white', activebackground='darkblue', activeforeground='white', padx=25, pady=10, relief='raised', text="Merge PDF Files", command=merge_redcliffe_pdf_files, font=("Arial", 12, "bold"), width=15)
merge_redcliffe_button2 = tk.Button(window, bg='magenta', fg='black', activebackground='gold', activeforeground='black', padx=25, pady=10, relief='raised', text="Merge All PDF Files", command=merge_all, font=("Arial", 12, "bold"), width=15)
merge_redcliffe_button1.place(x=615, y=58, anchor='ne')
merge_redcliffe_button2.place(x=613, y=178, anchor='ne')

pdf_rename_label = tk.Label(window, text="File Renaming System", font=("Arial", 16, "bold"))
pdf_rename_label.pack(pady=10, padx=37, anchor='w')

pdf_rename_button1 = tk.Button(window, bg='orange', fg='black', activebackground='darkblue', activeforeground='white', padx=25, pady=10, relief='raised', text="Rename PDF Files", command=rename_pdf_files, font=("Arial", 12, "bold"), width=15)
pdf_rename_button1.pack(pady=8, padx=45, anchor='w')

generate_excel_label = tk.Label(window, text="Count of Individual's Tests", font=("Arial", 16, "bold"))
generate_excel_label.place(x=305, y=130, anchor='ne')

generate_excel_button = tk.Button(window, bg='pink',fg='black', activebackground='darkblue', activeforeground='white',padx=25, pady=10, relief='raised', text="Patient's Test Count", command=count_of_tests_for_individual_patient, font=("Arial", 12, "bold"), width=15)
generate_excel_button.place(x=250, y=180, anchor='ne')

check_pdf_File = tk.Label(window, text="Check Pdf Files", font=("Arial", 16, "bold"))
check_pdf_File.place(x=930, y=10, anchor='ne')

check_pdf_button = tk.Button(window, bg='green',fg='black', activebackground='darkblue', activeforeground='white',padx=25, pady=10, relief='raised', text="Check Pdf Files", command=check_pdf_files, font=("Arial", 12, "bold"), width=15)
check_pdf_button.place(x=956, y=57, anchor='ne')

check_pdf_File = tk.Label(window, text="Split Pdf Files", font=("Arial", 16, "bold"))
check_pdf_File.place(x=903, y=130, anchor='ne')

check_pdf_button = tk.Button(window, bg='yellow',fg='black', activebackground='darkblue', activeforeground='white',padx=25, pady=10, relief='raised', text="Split Pdf Files", command=split_patient_file, font=("Arial", 12, "bold"), width=15)
check_pdf_button.place(x=955, y=175, anchor='ne')

# Label for the check Generate Excel for Merged Files button.
check_ecg_files_label = tk.Label(window, text="Data Extraction For Merged Files", font=("Arial", 16, "bold"))
check_ecg_files_label.place(x = 425, y=250, anchor='ne')
# Button for the check ecg file label.
check_ecg_files_button = tk.Button(window, bg='grey',fg='black', activebackground='darkgrey', activeforeground='white',padx=25, pady=10, relief='raised', text="Generate Excel for Merged Files", command=generate_excel_for_merged_files, font=("Arial", 12, "bold"), width=25)
check_ecg_files_button.place(x= 410, y=310, anchor='ne')

# Label for the check ecg file button.
check_ecg_files_label = tk.Label(window, text="Data Extraction For Individual Files", font=("Arial", 16, "bold"))
check_ecg_files_label.place(x = 860, y=250, anchor='ne')
# Button for the check ecg file label.
check_ecg_files_button = tk.Button(window, bg='red',fg='black', activebackground='red', activeforeground='white',padx=25, pady=10, relief='raised', text="Generate Excel For Individual File", command=generate_excel_for_individual_files, font=("Arial", 12, "bold"), width=25)
check_ecg_files_button.place(x=835, y=310, anchor='ne')


# dcm_to_pdf = tk.Label(window, text="Reports Observation", font=("Arial", 16, "bold"))
# dcm_to_pdf.place(x=233, y=255, anchor='ne')
#
# dcm_to_pdf_button = tk.Button(window, bg='red',fg='black', activebackground='darkblue', activeforeground='white',padx=30, pady=10, relief='raised', text="GET REPORTS OBSERVATION", command=dcm_to_pdf_converter, font=("Arial", 12, "bold"))
# dcm_to_pdf_button.place(x=328, y=300, anchor='ne')



window.mainloop()






